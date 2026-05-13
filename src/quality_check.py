"""
Pre-processing quality checks for raw input files.

Each check is a small function with signature:
    check(file_path, **context) -> (passed: bool, message: str)

`run_quality_checks` orchestrates them, writes a log file grouped by section,
and returns True if every check passed.
"""

import codecs
import os
import re
from datetime import datetime

import pandas as pd
from rdkit import Chem
from rdkit import RDLogger

# RDKit normally prints noisy errors on bad SMILES to stderr; we report them
# in the QC log instead, so silence its built-in logger here.
RDLogger.DisableLog("rdApp.*")


# ---------- Configuration ----------

MAX_FILE_SIZE_GB = 10
MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_GB * 1024 ** 3

MIN_BATCH_NUMBER = 0
MAX_BATCH_NUMBER = 10000

# Filename format: asms_<provider>_<batch>_<library>_<date>.csv
# Library names may contain underscores; date (YYYYMMDD) anchors the tail.
FILENAME_RE = re.compile(
    r"^asms_(?P<provider>[a-z]+)_(?P<batch>\d{1,5})_(?P<library>.+)_(?P<date>\d{8})\.csv$",
    re.IGNORECASE,
)

# Allowed characters in the filename (alphanumeric, underscore, period, hyphen).
FILENAME_ALLOWED_RE = re.compile(r"^[A-Za-z0-9_.\-]+$")


# ---------- Helpers ----------

def _parse_filename(file_path):
    """Returns the regex match object or None."""
    return FILENAME_RE.match(os.path.basename(file_path))


def _load_providers(providers_csv_path):
    """Load valid provider acronyms from a CSV with an `acronym` column.

    Returns a list of lowercase strings, or None if the file is missing.
    """
    if not providers_csv_path or not os.path.exists(providers_csv_path):
        return None
    try:
        df = pd.read_csv(providers_csv_path)
    except Exception:
        return None
    if "acronym" not in df.columns:
        return []
    return [str(a).strip().lower() for a in df["acronym"].dropna()]


def _load_data_generators(providers_csv_path):
    """Load registered data-generator names from Providers.csv.

    Reads the `data_generator_name` column when present. Returns a set of
    strings (exact match, no case-folding) or None if the file is missing.
    Returns an empty set if the column is absent (older Providers.csv files
    that haven't been extended yet).
    """
    if not providers_csv_path or not os.path.exists(providers_csv_path):
        return None
    try:
        df = pd.read_csv(providers_csv_path)
    except Exception:
        return None
    if "data_generator_name" not in df.columns:
        return set()
    return {str(g).strip() for g in df["data_generator_name"].dropna() if str(g).strip()}


def _list_libraries(masterlist_dir):
    """Return library names (filename stems) from MasterLists/, excluding the
    MasterList_Information mapping file. None if directory is missing.
    """
    if not masterlist_dir or not os.path.isdir(masterlist_dir):
        return None
    libs = []
    for name in os.listdir(masterlist_dir):
        if name == "MasterList_Information.xlsx":
            continue
        stem, ext = os.path.splitext(name)
        if ext.lower() in (".xlsx", ".xls", ".csv"):
            libs.append(stem)
    return libs


def _split_smiles(value):
    """Split a SMILES cell on ';' (the isomer separator) and trim each part.

    Empty parts are dropped. Used by SMILES validity and library-membership
    checks so a row with `'CC;CCC'` (a pre-Step-4 isomer group) is validated
    one component at a time rather than failing as a single malformed SMILES.
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return []
    return [p.strip() for p in str(value).split(";") if p.strip()]


def _load_associated_library_df(input_file_path, masterlist_dir, masterlist_info_path):
    """Resolve the master list file for `input_file_path` via MasterList_Information
    and return its full DataFrame, or None on any failure (missing folder,
    missing mapping file, mapping row absent, library file absent).

    Column existence is checked by the callers since different checks need
    different columns.
    """
    if not (masterlist_dir and os.path.isdir(masterlist_dir)):
        return None
    if not (masterlist_info_path and os.path.exists(masterlist_info_path)):
        return None
    try:
        info = pd.read_excel(masterlist_info_path)
    except Exception:
        return None
    if not {"FileName", "MaterListName"}.issubset(info.columns):
        return None
    file_name = os.path.basename(input_file_path)
    match = info.loc[info["FileName"] == file_name, "MaterListName"]
    if match.empty:
        return None
    lib_name = match.values[0]
    lib_path = os.path.join(masterlist_dir, f"{lib_name}.xlsx")
    if not os.path.exists(lib_path):
        return None
    try:
        return pd.read_excel(lib_path)
    except Exception:
        return None


def _load_associated_library_smiles(input_file_path, masterlist_dir, masterlist_info_path):
    """Returns a set of SMILES from the associated library, or None."""
    lib_df = _load_associated_library_df(input_file_path, masterlist_dir, masterlist_info_path)
    if lib_df is None or "SMILES" not in lib_df.columns:
        return None
    return set(lib_df["SMILES"].dropna().astype(str))


def _load_associated_library_formulas(input_file_path, masterlist_dir, masterlist_info_path):
    """Returns the set of formula strings from the associated library, or None.

    The library's formula column is matched case-insensitively (`formula`,
    `Formula`, `FORMULA` all work). Values are trimmed.
    """
    lib_df = _load_associated_library_df(input_file_path, masterlist_dir, masterlist_info_path)
    if lib_df is None:
        return None
    cols_lower = {c.lower(): c for c in lib_df.columns}
    if "formula" not in cols_lower:
        return None
    return set(lib_df[cols_lower["formula"]].dropna().astype(str).str.strip())


def _load_dataframe(file_path):
    """Read the input CSV once and drop fully duplicate rows.

    Column-content checks run against this cleaned DataFrame so that the
    duplicates flagged by the row-content check (Check 15) do not skew the
    per-column metrics. Returns None if the file cannot be parsed.
    """
    try:
        df = pd.read_csv(file_path)
    except Exception:
        return None
    return df.drop_duplicates(keep="first").reset_index(drop=True)


def _load_meta_columns(meta_csv_path):
    """Load the valid column names from the ASMS Meta Data reference CSV.

    The reference file's header row lists the canonical column names; the
    second row holds data types. Whitespace is stripped and duplicates
    (e.g. an accidental trailing-space variant) are collapsed.

    Returns a list of strings, or None if the file is missing/unreadable.
    """
    if not meta_csv_path or not os.path.exists(meta_csv_path):
        return None
    try:
        df = pd.read_csv(meta_csv_path, nrows=0)
    except Exception:
        return None
    seen = []
    for col in df.columns:
        name = str(col).strip()
        if name and name not in seen:
            seen.append(name)
    return seen


# ---------- File-format checks ----------

def check_file_opens(file_path, **_):
    """File can be opened for reading."""
    try:
        with open(file_path, "rb") as f:
            f.read(1)
    except OSError as e:
        return False, f"could not open file: {e}"
    return True, "file opens for reading"


def check_is_csv(file_path, **_):
    """File has a .csv extension."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return True, "extension is '.csv'"
    return False, f"expected '.csv', got '{ext or '(no extension)'}'"


def check_file_not_empty(file_path, **_):
    """File size is greater than zero bytes."""
    try:
        size = os.path.getsize(file_path)
    except OSError as e:
        return False, f"could not stat file: {e}"
    if size == 0:
        return False, "file is empty (0 bytes)"
    return True, f"file size is {size:,} bytes ({size / (1024 ** 2):.2f} MB)"


def check_file_size_under_limit(file_path, **_):
    """File size is below the configured limit (default 10 GB)."""
    try:
        size = os.path.getsize(file_path)
    except OSError as e:
        return False, f"could not stat file: {e}"
    if size > MAX_FILE_SIZE_BYTES:
        return False, (
            f"file size {size / (1024 ** 3):.2f} GB exceeds the "
            f"{MAX_FILE_SIZE_GB} GB limit"
        )
    return True, (
        f"file size {size / (1024 ** 3):.4f} GB is within the "
        f"{MAX_FILE_SIZE_GB} GB limit"
    )


def check_encoding_is_utf8(file_path, chunk_size=1024 * 1024, **_):
    """File contents decode cleanly as UTF-8 (reads the whole file in chunks)."""
    decoder = codecs.getincrementaldecoder("utf-8")()
    try:
        with open(file_path, "rb") as f:
            while True:
                chunk = f.read(chunk_size)
                if not chunk:
                    decoder.decode(b"", final=True)
                    break
                decoder.decode(chunk, final=False)
    except UnicodeDecodeError as e:
        return False, f"not UTF-8: {e}"
    except OSError as e:
        return False, f"read error: {e}"
    return True, "decodes as UTF-8"


def check_csv_parseable(file_path, **_):
    """File parses as CSV via pandas; also report rows, columns, and column names."""
    try:
        df = pd.read_csv(file_path, nrows=5)
    except Exception as e:
        return False, f"pandas could not parse as CSV: {e}"
    n_cols = len(df.columns)
    if n_cols < 1:
        return False, "parsed but has zero columns"

    # Count total data rows (line count minus header). Streamed so it stays
    # cheap even for multi-GB files.
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            line_count = sum(1 for _ in f)
        rows_msg = f"{max(line_count - 1, 0):,} rows"
    except OSError:
        rows_msg = "row count unavailable"

    col_list = ", ".join(df.columns)
    return True, (
        f"parsed as CSV with {n_cols} columns and {rows_msg}. "
        f"Columns: {col_list}"
    )


def check_columns_match_metadata(file_path, meta_columns=None, **_):
    """File's column names match the reference list in ASMS Meta Data.csv."""
    if meta_columns is None:
        return False, "metadata reference unavailable (ASMS Meta Data.csv not found)"
    if not meta_columns:
        return False, "metadata reference has no columns"

    try:
        df = pd.read_csv(file_path, nrows=0)
    except Exception as e:
        return False, f"could not read file columns: {e}"

    file_cols = [str(c).strip() for c in df.columns]
    valid = set(meta_columns)
    file_set = set(file_cols)

    missing = [c for c in meta_columns if c not in file_set]
    extra   = [c for c in file_cols    if c not in valid]

    if not missing and not extra:
        return True, "columns match ASMS Meta Data.csv reference"

    parts = ["columns do not match ASMS Meta Data.csv"]
    if missing:
        parts.append(f"missing required columns ({len(missing)}): {missing}")
    if extra:
        parts.append(f"extra columns not in reference ({len(extra)}): {extra}")
    return False, "; ".join(parts)


# ---------- Filename-format checks ----------

def check_filename_no_special_chars(file_path, **_):
    """Filename has only alphanumerics, underscores, periods, and hyphens."""
    name = os.path.basename(file_path)
    if FILENAME_ALLOWED_RE.match(name):
        return True, "no special characters or spaces in filename"
    bad = sorted(set(c for c in name if not re.match(r"[A-Za-z0-9_.\-]", c)))
    return False, f"filename contains disallowed character(s): {bad}"


def check_filename_starts_with_asms(file_path, **_):
    """Filename begins with 'asms_'."""
    name = os.path.basename(file_path)
    if name.lower().startswith("asms_"):
        return True, "filename starts with 'asms_'"
    return False, f"filename should start with 'asms_', got '{name[:10]}...'"


def check_filename_overall_format(file_path, **_):
    """Filename matches asms_<provider>_<batch>_<library>_<date>.csv."""
    match = _parse_filename(file_path)
    if match:
        return True, (
            f"parsed: provider='{match.group('provider')}', "
            f"batch='{match.group('batch')}', "
            f"library='{match.group('library')}', "
            f"date='{match.group('date')}'"
        )
    return False, (
        "filename does not match 'asms_<provider>_<batchN>_<library>_<YYYYMMDD>.csv'"
    )


def check_provider_acronym(file_path, providers=None, **_):
    """Provider acronym in the filename is in the registered list."""
    match = _parse_filename(file_path)
    if not match:
        return False, "filename did not parse; cannot extract provider"
    provider = match.group("provider").lower()

    if providers is None:
        return False, "providers list unavailable (Providers.csv not found)"
    if not providers:
        return False, "providers list is empty"
    if provider in providers:
        return True, f"provider '{provider}' is registered"
    return False, f"provider '{provider}' not in registered list: {providers}"


def check_batch_number_range(file_path, **_):
    """Batch number in the filename is between MIN_BATCH_NUMBER and MAX_BATCH_NUMBER."""
    match = _parse_filename(file_path)
    if not match:
        return False, "filename did not parse; cannot extract batch number"
    batch_str = match.group("batch")
    try:
        batch_int = int(batch_str)
    except ValueError:
        return False, f"batch number '{batch_str}' is not an integer"
    if MIN_BATCH_NUMBER <= batch_int <= MAX_BATCH_NUMBER:
        return True, f"batch number {batch_int} (from '{batch_str}') is in range [{MIN_BATCH_NUMBER}, {MAX_BATCH_NUMBER}]"
    return False, (
        f"batch number {batch_int} is outside [{MIN_BATCH_NUMBER}, {MAX_BATCH_NUMBER}]"
    )


def check_library_name(file_path, libraries=None, **_):
    """Library name in the filename matches a file in MasterLists/."""
    match = _parse_filename(file_path)
    if not match:
        return False, "filename did not parse; cannot extract library name"
    library = match.group("library")

    if libraries is None:
        return False, "libraries list unavailable (MasterLists/ not found)"
    if not libraries:
        return False, "no registered libraries (MasterLists/ is empty)"
    if library in libraries:
        return True, f"library '{library}' is registered"
    return False, f"library '{library}' not in registered list: {libraries}"


def check_date_valid_and_not_future(file_path, **_):
    """Date in the filename is a valid YYYYMMDD and not in the future."""
    match = _parse_filename(file_path)
    if not match:
        return False, "filename did not parse; cannot extract date"
    date_str = match.group("date")
    try:
        date_obj = datetime.strptime(date_str, "%Y%m%d").date()
    except ValueError as e:
        return False, f"date '{date_str}' is not a valid YYYYMMDD: {e}"
    today = datetime.now().date()
    if date_obj > today:
        return False, f"date {date_obj.isoformat()} is in the future (today is {today.isoformat()})"
    return True, f"date {date_obj.isoformat()} is valid and not in the future"


# ---------- Row content checks ----------

def check_no_duplicate_rows(file_path, output_dir=None, **_):
    """No two rows in the file are identical across all columns.

    Emits a WARN (not FAIL) when duplicates exist, because Step 3
    (anomaly_selection) will drop them anyway. Also writes
    `FullyDuplicate_rows_report.csv` into `output_dir` listing every row that
    is part of a duplicate group (all copies, not just the dropped ones).
    """
    try:
        df = pd.read_csv(file_path)
    except Exception as e:
        return False, f"could not read file: {e}"

    dup_mask_first = df.duplicated(keep="first")
    n_dups = int(dup_mask_first.sum())
    if n_dups == 0:
        return True, f"no fully duplicate rows (checked {len(df):,} rows)"

    # Write a report of every row in any duplicate group (keep=False shows all copies)
    report_msg = ""
    if output_dir:
        try:
            os.makedirs(output_dir, exist_ok=True)
            report_df = df[df.duplicated(keep=False)].copy()
            report_df.insert(0, "FileLine", report_df.index + 2)  # +2: 1-index + header
            report_path = os.path.join(output_dir, "FullyDuplicate_rows_report.csv")
            report_df.to_csv(report_path, index=False)
            report_msg = f"; report saved to {report_path}"
        except Exception as e:
            report_msg = f"; failed to write report: {e}"

    # 1-indexed line numbers including the header row (so row 0 in df is line 2)
    dup_line_nums = [int(i) + 2 for i in dup_mask_first[dup_mask_first].index[:5]]
    suffix = f" (and {n_dups - 5} more)" if n_dups > 5 else ""
    return True, (
        f"found {n_dups:,} fully duplicate row(s) "
        f"(extras will be removed in Step 3). "
        f"First duplicates at file line(s): {dup_line_nums}{suffix}{report_msg}"
    ), "WARN"


# ---------- Column content check helpers ----------
#
# All column-content checks operate on the DataFrame pre-loaded into the
# orchestrator context (`df`), which has been read with pandas.read_csv and
# fully-duplicate-row-deduped. Each helper handles the standard preconditions
# (df present, column present) so the per-column wrapper stays one line.

def _ensure_df_and_columns(df, *columns):
    """Returns (ok, fail_message). ok=True means all required columns exist."""
    if df is None:
        return False, "dataframe unavailable (file could not be read)"
    missing = [c for c in columns if c not in df.columns]
    if missing:
        return False, f"required column(s) missing from file: {missing}"
    return True, None


def _check_column_is_string(df, column):
    """Column dtype is object and every non-null value is a Python str."""
    ok, fail = _ensure_df_and_columns(df, column)
    if not ok:
        return False, fail
    col = df[column]
    non_null = col.dropna()
    if non_null.empty:
        return True, f"'{column}' is empty (no values to type-check)"
    bad = sorted({type(v).__name__ for v in non_null if not isinstance(v, str)})
    if bad:
        return False, (
            f"'{column}' has {len(bad)} non-string type(s): {bad} "
            f"(pandas dtype: {col.dtype})"
        )
    return True, f"all {len(non_null):,} non-null '{column}' values are strings"


def _check_column_no_whitespace(df, column):
    """No values in the column have leading or trailing whitespace."""
    ok, fail = _ensure_df_and_columns(df, column)
    if not ok:
        return False, fail
    col = df[column].dropna().astype(str)
    has_ws = col.str.startswith(" ") | col.str.endswith(" ") | col.str.startswith("\t") | col.str.endswith("\t")
    n = int(has_ws.sum())
    if n == 0:
        return True, f"no leading/trailing whitespace in '{column}'"
    return False, f"{n:,} '{column}' value(s) have leading/trailing whitespace"


def _check_column_is_int(df, column):
    """Column dtype is integer-like, or all non-null values are whole numbers."""
    ok, fail = _ensure_df_and_columns(df, column)
    if not ok:
        return False, fail
    col = df[column]
    non_null = col.dropna()
    if non_null.empty:
        return True, f"'{column}' is empty (no values to type-check)"
    if pd.api.types.is_integer_dtype(col):
        return True, f"all {len(non_null):,} '{column}' values are integers (dtype: {col.dtype})"
    # Accept float dtype only if every value happens to be a whole number,
    # which is what pandas gives you for an integer column that contains NaN.
    if pd.api.types.is_float_dtype(col):
        whole = non_null == non_null.astype("int64", errors="ignore")
        try:
            if whole.all():
                return True, (
                    f"'{column}' is float dtype ({col.dtype}) but all "
                    f"{len(non_null):,} values are whole numbers"
                )
        except Exception:
            pass
        return False, (
            f"'{column}' has non-integer numeric values (dtype: {col.dtype})"
        )
    return False, f"'{column}' is not integer (dtype: {col.dtype})"


def _check_column_is_numeric(df, column):
    """Column dtype is numeric (int or float), or every value coerces cleanly to a number."""
    ok, fail = _ensure_df_and_columns(df, column)
    if not ok:
        return False, fail
    col = df[column]
    non_null = col.dropna()
    if non_null.empty:
        return True, f"'{column}' is empty (no values to type-check)"
    if pd.api.types.is_numeric_dtype(col):
        return True, (
            f"all {len(non_null):,} '{column}' values are numeric (dtype: {col.dtype})"
        )
    coerced = pd.to_numeric(col, errors="coerce")
    n_unparseable = int((coerced.isna() & col.notna()).sum())
    if n_unparseable:
        return False, (
            f"'{column}' is not numeric (dtype: {col.dtype}); "
            f"{n_unparseable:,} value(s) cannot be parsed as numbers"
        )
    return True, f"'{column}' is parseable as numbers (stored as {col.dtype})"


def _check_column_positive(df, column):
    """All non-null values in `column` are strictly positive (> 0)."""
    ok, fail = _ensure_df_and_columns(df, column)
    if not ok:
        return False, fail
    coerced = pd.to_numeric(df[column], errors="coerce")
    non_positive = coerced[coerced <= 0]
    n_unparseable = int(coerced.isna().sum() - df[column].isna().sum())
    if non_positive.empty and n_unparseable == 0:
        return True, f"all '{column}' values are positive (> 0)"
    parts = []
    if not non_positive.empty:
        sample = non_positive.dropna().unique()[:5].tolist()
        parts.append(f"{len(non_positive):,} value(s) <= 0 (e.g. {sample})")
    if n_unparseable:
        parts.append(f"{n_unparseable:,} value(s) could not be parsed as numbers")
    return False, f"'{column}': " + "; ".join(parts)


def _check_column_is_bool(df, column):
    """Column dtype is boolean."""
    ok, fail = _ensure_df_and_columns(df, column)
    if not ok:
        return False, fail
    col = df[column]
    non_null = col.dropna()
    if non_null.empty:
        return True, f"'{column}' is empty (no values to type-check)"
    if pd.api.types.is_bool_dtype(col):
        return True, f"all {len(non_null):,} '{column}' values are booleans (dtype: {col.dtype})"
    types = sorted({type(v).__name__ for v in non_null})
    return False, (
        f"'{column}' is not boolean (dtype: {col.dtype}, value types: {types})"
    )


def _check_column_in_set(df, column, allowed):
    """All non-null values in `column` belong to `allowed` (a set or iterable)."""
    ok, fail = _ensure_df_and_columns(df, column)
    if not ok:
        return False, fail
    col = df[column]
    non_null = col.dropna()
    if non_null.empty:
        return True, f"'{column}' is empty"
    allowed_set = set(allowed)
    bad_mask = ~non_null.isin(allowed_set)
    n_bad = int(bad_mask.sum())
    if n_bad == 0:
        return True, f"all '{column}' values are in {sorted(allowed_set, key=str)}"
    sample = non_null[bad_mask].unique()[:5].tolist()
    return False, (
        f"{n_bad:,} '{column}' value(s) not in {sorted(allowed_set, key=str)} "
        f"(e.g. {sample})"
    )


def _check_column_at_least(df, column, threshold):
    """All non-null values in `column` are >= `threshold`."""
    ok, fail = _ensure_df_and_columns(df, column)
    if not ok:
        return False, fail
    coerced = pd.to_numeric(df[column], errors="coerce")
    below = coerced[coerced < threshold]
    n_unparseable = int((coerced.isna() & df[column].notna()).sum())
    if below.empty and n_unparseable == 0:
        return True, f"all '{column}' values are >= {threshold:,}"
    parts = []
    if not below.empty:
        sample = below.dropna().unique()[:5].tolist()
        parts.append(f"{len(below):,} value(s) < {threshold:,} (e.g. {sample})")
    if n_unparseable:
        parts.append(f"{n_unparseable:,} value(s) could not be parsed as numbers")
    return False, f"'{column}': " + "; ".join(parts)


def _check_column_equals(df, column, expected, atol=1e-9):
    """All non-null values in `column` equal `expected` (within float tolerance)."""
    ok, fail = _ensure_df_and_columns(df, column)
    if not ok:
        return False, fail
    coerced = pd.to_numeric(df[column], errors="coerce")
    n_unparseable = int((coerced.isna() & df[column].notna()).sum())
    # Compare only the rows that successfully coerced
    valid = coerced.dropna()
    not_equal = valid[~(valid.sub(expected).abs() <= atol)]
    if not_equal.empty and n_unparseable == 0:
        return True, f"all '{column}' values equal {expected}"
    parts = []
    if not not_equal.empty:
        sample = not_equal.unique()[:5].tolist()
        parts.append(f"{len(not_equal):,} value(s) != {expected} (e.g. {sample})")
    if n_unparseable:
        parts.append(f"{n_unparseable:,} value(s) could not be parsed as numbers")
    return False, f"'{column}': " + "; ".join(parts)


def _check_column_in_range(df, column, lo, hi, lo_inclusive=True, hi_inclusive=True):
    """All non-null values in `column` are within the range between `lo` and `hi`.

    By default bounds are inclusive on both ends (`[lo, hi]`). Pass
    `lo_inclusive=False` / `hi_inclusive=False` to use open bounds (e.g.
    `(0, 6)` to require strictly `0 < value < 6`).
    """
    ok, fail = _ensure_df_and_columns(df, column)
    if not ok:
        return False, fail
    coerced = pd.to_numeric(df[column], errors="coerce")
    n_unparseable = int((coerced.isna() & df[column].notna()).sum())
    below = coerced < lo if lo_inclusive else coerced <= lo
    above = coerced > hi if hi_inclusive else coerced >= hi
    out_of_range = coerced[below | above]
    n_out = int(out_of_range.shape[0])
    range_str = (
        f"{'[' if lo_inclusive else '('}{lo}, {hi}{']' if hi_inclusive else ')'}"
    )
    if n_out == 0 and n_unparseable == 0:
        return True, f"all '{column}' values are in {range_str}"
    parts = []
    if n_out:
        sample = out_of_range.dropna().unique()[:5].tolist()
        parts.append(f"{n_out:,} value(s) outside {range_str} (e.g. {sample})")
    if n_unparseable:
        parts.append(f"{n_unparseable:,} value(s) could not be parsed as numbers")
    return False, f"'{column}': " + "; ".join(parts)


def _check_column_min_length(df, column, min_length):
    """All non-null values in `column` have length strictly greater than `min_length`."""
    ok, fail = _ensure_df_and_columns(df, column)
    if not ok:
        return False, fail
    col = df[column].dropna().astype(str)
    too_short = col[col.str.len() <= min_length]
    n = int(too_short.shape[0])
    if n == 0:
        return True, f"all '{column}' values have length > {min_length}"
    sample = too_short.unique()[:5].tolist()
    return False, (
        f"{n:,} '{column}' value(s) have length <= {min_length} (e.g. {sample})"
    )


def _check_column_no_nulls(df, column):
    """No null/NaN values in the column."""
    ok, fail = _ensure_df_and_columns(df, column)
    if not ok:
        return False, fail
    n = int(df[column].isna().sum())
    if n == 0:
        return True, f"no null values in '{column}'"
    return False, f"{n:,} null value(s) in '{column}'"


def _check_no_duplicates_per_group(df, column, group_col, output_dir):
    """Within each `group_col` value, no two rows share the same `column` value.

    Writes one CSV report per offending `group_col` value (sanitized filename),
    listing every offending row. Emits WARN rather than FAIL.
    """
    ok, fail = _ensure_df_and_columns(df, column, group_col)
    if not ok:
        return False, fail
    dup_mask = df.duplicated(subset=[group_col, column], keep=False)
    if not dup_mask.any():
        return True, f"no duplicate '{column}' values within any '{group_col}' group"

    n_dup_rows = int(dup_mask.sum())
    dup_df = df[dup_mask].copy()
    reports = []
    if output_dir:
        try:
            os.makedirs(output_dir, exist_ok=True)
            for group_val, group_rows in dup_df.groupby(group_col):
                safe = re.sub(r"[^A-Za-z0-9_.\-]", "_", str(group_val))
                report_path = os.path.join(
                    output_dir, f"duplicate_{column}_per_{group_col}_{safe}.csv"
                )
                group_rows.to_csv(report_path, index=False)
                reports.append(os.path.basename(report_path))
        except Exception as e:
            return True, (
                f"found {n_dup_rows:,} duplicate '{column}' rows across "
                f"{dup_df[group_col].nunique()} '{group_col}' group(s); "
                f"failed to write reports: {e}"
            ), "WARN"

    n_groups = dup_df[group_col].nunique()
    return True, (
        f"found {n_dup_rows:,} duplicate '{column}' row(s) across "
        f"{n_groups} '{group_col}' group(s); reports: {reports}"
    ), "WARN"


# ---------- COMPOUND_ID checks ----------

def check_compound_id_is_string(file_path, df=None, **_):
    """COMPOUND_ID dtype is string (VARCHAR-equivalent)."""
    return _check_column_is_string(df, "COMPOUND_ID")


def check_compound_id_no_whitespace(file_path, df=None, **_):
    """COMPOUND_ID values have no leading/trailing whitespace."""
    return _check_column_no_whitespace(df, "COMPOUND_ID")


def check_compound_id_no_nulls(file_path, df=None, **_):
    """COMPOUND_ID has no null values."""
    return _check_column_no_nulls(df, "COMPOUND_ID")


def check_compound_id_unique_per_target(file_path, df=None, output_dir=None, **_):
    """Each COMPOUND_ID appears at most once per TARGET_ID."""
    return _check_no_duplicates_per_group(df, "COMPOUND_ID", "TARGET_ID", output_dir)


# ---------- SMILES checks ----------

def check_smiles_is_string(file_path, df=None, **_):
    """SMILES dtype is string (VARCHAR-equivalent)."""
    return _check_column_is_string(df, "SMILES")


def check_smiles_no_whitespace(file_path, df=None, **_):
    """SMILES values have no leading/trailing whitespace."""
    return _check_column_no_whitespace(df, "SMILES")


def check_smiles_no_nulls(file_path, df=None, **_):
    """SMILES has no null values."""
    return _check_column_no_nulls(df, "SMILES")


def check_smiles_unique_per_target(file_path, df=None, output_dir=None, **_):
    """Each SMILES appears at most once per TARGET_ID."""
    return _check_no_duplicates_per_group(df, "SMILES", "TARGET_ID", output_dir)


def check_smiles_valid(file_path, df=None, output_dir=None, **_):
    """Every SMILES is non-empty and RDKit-parseable.

    Isomer groups (parts separated by ';') are split and each component is
    validated independently, so a pre-Step-4 row like `'CC;CCC'` is checked
    as two valid molecules rather than failing as one malformed SMILES.

    Writes `invalid_smiles_report.csv` listing every offending row and the
    type of problem (`empty` or `malformed: '<part>'`).
    """
    ok, fail = _ensure_df_and_columns(df, "SMILES")
    if not ok:
        return False, fail

    smiles_col = df["SMILES"]

    # Identify problems per row. Build maps so we only call MolFromSmiles
    # once per unique part across the whole column.
    row_parts = {}
    unique_parts = set()
    for idx, raw in smiles_col.items():
        parts = _split_smiles(raw)
        row_parts[idx] = parts
        unique_parts.update(parts)

    invalid_parts = {p for p in unique_parts if Chem.MolFromSmiles(p) is None}

    empty_rows = []
    malformed_rows = []   # list of (idx, first_invalid_part)
    for idx, parts in row_parts.items():
        if not parts:
            empty_rows.append(idx)
        else:
            bad = next((p for p in parts if p in invalid_parts), None)
            if bad is not None:
                malformed_rows.append((idx, bad))

    n_empty = len(empty_rows)
    n_malformed = len(malformed_rows)
    if n_empty == 0 and n_malformed == 0:
        return True, f"all {len(df):,} SMILES rows are valid (non-empty, RDKit-parseable)"

    report_msg = ""
    if output_dir:
        try:
            os.makedirs(output_dir, exist_ok=True)
            offending_indices = sorted(set(empty_rows) | {i for i, _ in malformed_rows})
            bad_part_for = {i: p for i, p in malformed_rows}
            report_df = df.loc[offending_indices].copy()
            report_df.insert(
                0, "Issue",
                ["empty" if i in set(empty_rows) else f"malformed: '{bad_part_for[i]}'"
                 for i in offending_indices],
            )
            report_path = os.path.join(output_dir, "invalid_smiles_report.csv")
            report_df.to_csv(report_path, index=False)
            report_msg = f"; report saved to {os.path.basename(report_path)}"
        except Exception as e:
            report_msg = f"; failed to write report: {e}"

    return False, (
        f"found {n_empty:,} empty and {n_malformed:,} malformed SMILES row(s){report_msg}"
    )


def check_smiles_in_library(file_path, df=None, masterlist_dir=None, output_dir=None, **_):
    """Every SMILES in the input is present in the master list referenced by
    `MasterList_Information.xlsx` for this raw CSV.

    Isomer groups are split on ';' and each component is checked against the
    library individually. WARN (not FAIL): a SMILES that's not in the
    declared library is a data-integrity flag worth surfacing, but does not
    block downstream processing.
    """
    ok, fail = _ensure_df_and_columns(df, "SMILES")
    if not ok:
        return False, fail
    if not masterlist_dir:
        return False, "masterlist_dir not provided; cannot check library membership"

    info_path = os.path.join(masterlist_dir, "MasterList_Information.xlsx")
    library_smiles = _load_associated_library_smiles(file_path, masterlist_dir, info_path)
    if library_smiles is None:
        return False, (
            "could not load associated library (check MasterList_Information.xlsx "
            "FileName/MaterListName mapping and the referenced library file)"
        )

    smiles_col = df["SMILES"]
    missing_rows = []   # list of (idx, first_missing_part)
    for idx, raw in smiles_col.items():
        for part in _split_smiles(raw):
            if part not in library_smiles:
                missing_rows.append((idx, part))
                break

    if not missing_rows:
        return True, (
            f"all SMILES are in the associated library "
            f"({len(library_smiles):,} library SMILES)"
        )

    report_msg = ""
    if output_dir:
        try:
            os.makedirs(output_dir, exist_ok=True)
            offending_indices = sorted({i for i, _ in missing_rows})
            missing_part_for = {i: p for i, p in missing_rows}
            report_df = df.loc[offending_indices].copy()
            report_df.insert(0, "MissingPart", [missing_part_for[i] for i in offending_indices])
            report_path = os.path.join(output_dir, "smiles_not_in_library_report.csv")
            report_df.to_csv(report_path, index=False)
            report_msg = f"; report saved to {os.path.basename(report_path)}"
        except Exception as e:
            report_msg = f"; failed to write report: {e}"

    return True, (
        f"found {len(missing_rows):,} row(s) with SMILES not in the associated library{report_msg}"
    ), "WARN"


# ---------- ASMS_BATCH_NAME checks ----------

# Matches "<provider>_<batchnumber>", e.g. "sgcto_01".  Provider is alphabetic,
# batch is digits. Case-insensitive; the actual provider is validated against
# Providers.csv at check time.
_BATCH_NAME_RE = re.compile(r"^([a-zA-Z]+)_(\d+)$")


def check_asms_batch_name_is_string(file_path, df=None, **_):
    """ASMS_BATCH_NAME dtype is string (VARCHAR-equivalent)."""
    return _check_column_is_string(df, "ASMS_BATCH_NAME")


def check_asms_batch_name_no_whitespace(file_path, df=None, **_):
    """ASMS_BATCH_NAME values have no leading/trailing whitespace."""
    return _check_column_no_whitespace(df, "ASMS_BATCH_NAME")


def check_asms_batch_name_no_nulls(file_path, df=None, **_):
    """ASMS_BATCH_NAME has no null values."""
    return _check_column_no_nulls(df, "ASMS_BATCH_NAME")


def check_asms_batch_name_consistent(file_path, df=None, **_):
    """All rows share the same ASMS_BATCH_NAME value (one file = one batch)."""
    ok, fail = _ensure_df_and_columns(df, "ASMS_BATCH_NAME")
    if not ok:
        return False, fail
    non_null = df["ASMS_BATCH_NAME"].dropna()
    if non_null.empty:
        return False, "ASMS_BATCH_NAME column has no non-null values"
    unique = non_null.unique()
    if len(unique) == 1:
        return True, f"all rows share ASMS_BATCH_NAME = '{unique[0]}'"
    preview = list(unique[:5])
    suffix = f" (and {len(unique) - 5} more)" if len(unique) > 5 else ""
    return False, (
        f"ASMS_BATCH_NAME is not consistent: found {len(unique)} distinct values: "
        f"{preview}{suffix}"
    )


# ---------- COMPOUND_FORMULA checks ----------

def check_compound_formula_is_string(file_path, df=None, **_):
    """COMPOUND_FORMULA dtype is string (VARCHAR-equivalent)."""
    return _check_column_is_string(df, "COMPOUND_FORMULA")


def check_compound_formula_no_whitespace(file_path, df=None, **_):
    """COMPOUND_FORMULA values have no leading/trailing whitespace."""
    return _check_column_no_whitespace(df, "COMPOUND_FORMULA")


def check_compound_formula_no_nulls(file_path, df=None, **_):
    """COMPOUND_FORMULA has no null values."""
    return _check_column_no_nulls(df, "COMPOUND_FORMULA")


def check_compound_formula_in_library(file_path, df=None, masterlist_dir=None,
                                      output_dir=None, **_):
    """Every COMPOUND_FORMULA value (or isomer component) appears in the
    `formula` column of the associated library.

    Isomer rows (';' separated formulas) are split and each component is
    checked independently. WARN: an unknown formula is a data-integrity flag
    worth reporting but does not block downstream processing.
    """
    ok, fail = _ensure_df_and_columns(df, "COMPOUND_FORMULA")
    if not ok:
        return False, fail
    if not masterlist_dir:
        return False, "masterlist_dir not provided; cannot check formula library membership"

    info_path = os.path.join(masterlist_dir, "MasterList_Information.xlsx")
    library_formulas = _load_associated_library_formulas(file_path, masterlist_dir, info_path)
    if library_formulas is None:
        return False, (
            "could not load associated library or its 'formula' column "
            "(check MasterList_Information.xlsx and that the library file "
            "has a `formula` column)"
        )

    missing_rows = []   # list of (row_idx, first_missing_part)
    for idx, raw in df["COMPOUND_FORMULA"].items():
        for part in _split_smiles(raw):
            if part not in library_formulas:
                missing_rows.append((idx, part))
                break

    if not missing_rows:
        return True, (
            f"all COMPOUND_FORMULA values are in the associated library "
            f"({len(library_formulas):,} library formulas)"
        )

    report_msg = ""
    if output_dir:
        try:
            os.makedirs(output_dir, exist_ok=True)
            offending = sorted({i for i, _ in missing_rows})
            missing_part_for = {i: p for i, p in missing_rows}
            report_df = df.loc[offending].copy()
            report_df.insert(0, "MissingFormula", [missing_part_for[i] for i in offending])
            report_path = os.path.join(output_dir, "formula_not_in_library_report.csv")
            report_df.to_csv(report_path, index=False)
            report_msg = f"; report saved to {os.path.basename(report_path)}"
        except Exception as e:
            report_msg = f"; failed to write report: {e}"

    return True, (
        f"found {len(missing_rows):,} row(s) with COMPOUND_FORMULA not in the associated library{report_msg}"
    ), "WARN"


# ---------- POOL_NAME checks ----------

def check_pool_name_is_string(file_path, df=None, **_):
    """POOL_NAME dtype is string (VARCHAR-equivalent)."""
    return _check_column_is_string(df, "POOL_NAME")


def check_pool_name_no_whitespace(file_path, df=None, **_):
    """POOL_NAME values have no leading/trailing whitespace."""
    return _check_column_no_whitespace(df, "POOL_NAME")


def check_pool_name_no_nulls(file_path, df=None, **_):
    """POOL_NAME has no null values."""
    return _check_column_no_nulls(df, "POOL_NAME")


# ---------- POOL_ID checks ----------

def check_pool_id_is_string(file_path, df=None, **_):
    """POOL_ID dtype is string (VARCHAR-equivalent)."""
    return _check_column_is_string(df, "POOL_ID")


def check_pool_id_no_whitespace(file_path, df=None, **_):
    """POOL_ID values have no leading/trailing whitespace."""
    return _check_column_no_whitespace(df, "POOL_ID")


def check_pool_id_no_nulls(file_path, df=None, **_):
    """POOL_ID has no null values."""
    return _check_column_no_nulls(df, "POOL_ID")


# ---------- TARGET_ID checks ----------

# Format: <name>_<UniprotID>_<startAA>_<endAA>
# `name` and `UniprotID` are alphanumeric; start/end are positive integers.
# The Uniprot_ID segment is currently only checked structurally — see the
# TODO at the bottom of this block for plans to validate it against a
# registered list when that list becomes available.
TARGET_ID_RE = re.compile(r"^([A-Za-z0-9]+)_([A-Za-z0-9]+)_(\d+)_(\d+)$")


def check_target_id_is_string(file_path, df=None, **_):
    """TARGET_ID dtype is string (VARCHAR-equivalent)."""
    return _check_column_is_string(df, "TARGET_ID")


def check_target_id_no_whitespace(file_path, df=None, **_):
    """TARGET_ID values have no leading/trailing whitespace."""
    return _check_column_no_whitespace(df, "TARGET_ID")


def check_target_id_no_nulls(file_path, df=None, **_):
    """TARGET_ID has no null values."""
    return _check_column_no_nulls(df, "TARGET_ID")


def check_target_id_format(file_path, df=None, **_):
    """TARGET_ID values match '<name>_<UniprotID>_<startAA>_<endAA>'."""
    ok, fail = _ensure_df_and_columns(df, "TARGET_ID")
    if not ok:
        return False, fail
    invalid = []
    for val in df["TARGET_ID"].dropna().unique():
        if not TARGET_ID_RE.match(str(val).strip()):
            invalid.append(str(val))
    if not invalid:
        return True, "all TARGET_ID values match '<name>_<UniprotID>_<startAA>_<endAA>'"
    preview = invalid[:5]
    suffix = f" (and {len(invalid) - 5} more)" if len(invalid) > 5 else ""
    return False, f"invalid TARGET_ID value(s): {preview}{suffix}"


def check_target_id_start_lt_end(file_path, df=None, **_):
    """The two trailing numeric segments of TARGET_ID satisfy start < end."""
    ok, fail = _ensure_df_and_columns(df, "TARGET_ID")
    if not ok:
        return False, fail
    bad = []   # list of (target_id, start, end)
    for val in df["TARGET_ID"].dropna().unique():
        m = TARGET_ID_RE.match(str(val).strip())
        if not m:
            continue   # caught by format check
        start, end = int(m.group(3)), int(m.group(4))
        if start >= end:
            bad.append((str(val), start, end))
    if not bad:
        return True, "all TARGET_IDs have start < end"
    preview = [f"'{s}' (start={a}, end={b})" for s, a, b in bad[:5]]
    suffix = f" (and {len(bad) - 5} more)" if len(bad) > 5 else ""
    return False, f"TARGET_ID(s) with start >= end: {preview}{suffix}"


def check_target_id_same_compound_count(file_path, df=None, **_):
    """Each TARGET_ID has the same number of unique COMPOUND_IDs.

    A batch is expected to test the same library against each target, so
    every TARGET_ID should appear with the same compound-count. Surfaces as
    WARN (not FAIL): differing counts can be intentional in some workflows.
    """
    ok, fail = _ensure_df_and_columns(df, "TARGET_ID", "COMPOUND_ID")
    if not ok:
        return False, fail
    counts = df.groupby("TARGET_ID")["COMPOUND_ID"].nunique()
    if counts.empty:
        return True, "TARGET_ID column is empty"
    unique_counts = sorted(counts.unique().tolist())
    if len(unique_counts) == 1:
        return True, (
            f"all {len(counts)} TARGET_IDs have {unique_counts[0]:,} unique COMPOUND_IDs"
        )
    return True, (
        f"TARGET_IDs have differing compound-counts: "
        f"min={int(counts.min()):,}, max={int(counts.max()):,}, "
        f"distinct counts seen: {unique_counts}"
    ), "WARN"


# TODO: Uniprot_ID validation against a registry.
# When a list of valid Uniprot_IDs becomes available, add a check here that
# pulls the UniprotID segment from the regex (group 2) and verifies it
# against the registry. Suggested wiring:
#   1. Load the list in the orchestrator and pass it via context as
#      `uniprot_ids` (similar to `providers` / `libraries`).
#   2. Implement check_target_id_uniprot_registered(file_path, df=None,
#      uniprot_ids=None, **_) that iterates unique TARGET_IDs and reports
#      any whose Uniprot_ID segment is not in the registry.
#   3. Register it in SECTIONS.


# ---------- PROTEIN_ID checks ----------

def check_protein_id_is_string(file_path, df=None, **_):
    """PROTEIN_ID dtype is string (VARCHAR-equivalent)."""
    return _check_column_is_string(df, "PROTEIN_ID")


def check_protein_id_no_whitespace(file_path, df=None, **_):
    """PROTEIN_ID values have no leading/trailing whitespace."""
    return _check_column_no_whitespace(df, "PROTEIN_ID")


def check_protein_id_no_nulls(file_path, df=None, **_):
    """PROTEIN_ID has no null values."""
    return _check_column_no_nulls(df, "PROTEIN_ID")


# ---------- INCUBATION_VOLUME checks ----------

# TODO: realistic experimental range check, to be implemented later.
# When the realistic [min, max] uL range is decided, set the constants and
# enable the range check below.
# INCUBATION_VOLUME_MIN = ?     # uL (e.g. 1)
# INCUBATION_VOLUME_MAX = ?     # uL (e.g. 1000)
#
# def check_incubation_volume_in_realistic_range(file_path, df=None, **_):
#     """INCUBATION_VOLUME (uL) is within the expected experimental range."""
#     return _check_column_in_range(
#         df, "INCUBATION_VOLUME",
#         INCUBATION_VOLUME_MIN, INCUBATION_VOLUME_MAX,
#     )
#
# Then register it in SECTIONS alongside the other INCUBATION_VOLUME checks.


# ---------- PROTEIN_CONC checks ----------

# Expected protein concentration value (uM) — change if the experimental
# protocol uses a different fixed concentration.
PROTEIN_CONC_EXPECTED = 1.0

# TODO: realistic experimental range check, to be implemented later.
# When the realistic [min, max] uM range is decided, set the constants and
# enable the range check below.
# PROTEIN_CONC_MIN = ?     # uM
# PROTEIN_CONC_MAX = ?     # uM
#
# def check_protein_conc_in_realistic_range(file_path, df=None, **_):
#     """PROTEIN_CONC (uM) is within the expected experimental range."""
#     return _check_column_in_range(
#         df, "PROTEIN_CONC",
#         PROTEIN_CONC_MIN, PROTEIN_CONC_MAX,
#     )
#
# Then register it in SECTIONS alongside the other PROTEIN_CONC checks.


# ---------- MS_REPRODUCABILITY checks ----------

# ---------- PROTEIN_TAG checks ----------

def check_protein_tag_is_string(file_path, df=None, **_):
    """PROTEIN_TAG dtype is string (VARCHAR-equivalent)."""
    return _check_column_is_string(df, "PROTEIN_TAG")


def check_protein_tag_no_whitespace(file_path, df=None, **_):
    """PROTEIN_TAG values have no leading/trailing whitespace."""
    return _check_column_no_whitespace(df, "PROTEIN_TAG")


def check_protein_tag_no_nulls(file_path, df=None, **_):
    """PROTEIN_TAG has no null values."""
    return _check_column_no_nulls(df, "PROTEIN_TAG")


# ---------- PROTEIN_SEQ checks ----------

# Minimum acceptable sequence length (strictly greater than this).
PROTEIN_SEQ_MIN_LENGTH = 6


def check_protein_seq_is_string(file_path, df=None, **_):
    """PROTEIN_SEQ dtype is string (VARCHAR-equivalent)."""
    return _check_column_is_string(df, "PROTEIN_SEQ")


def check_protein_seq_no_whitespace(file_path, df=None, **_):
    """PROTEIN_SEQ values have no leading/trailing whitespace."""
    return _check_column_no_whitespace(df, "PROTEIN_SEQ")


def check_protein_seq_no_nulls(file_path, df=None, **_):
    """PROTEIN_SEQ has no null values."""
    return _check_column_no_nulls(df, "PROTEIN_SEQ")


def check_protein_seq_min_length(file_path, df=None, **_):
    """PROTEIN_SEQ values are longer than PROTEIN_SEQ_MIN_LENGTH characters."""
    return _check_column_min_length(df, "PROTEIN_SEQ", PROTEIN_SEQ_MIN_LENGTH)


# ---------- CHIRAL_SELECTIVITY checks ----------

# Allowed values for CHIRAL_SELECTIVITY (case-sensitive).
CHIRAL_SELECTIVITY_ALLOWED = {
    "achiral",
    "chiral_selective",
    "chiral_not_selective",
    "chiral_undetermined",
}


def check_chiral_selectivity_is_string(file_path, df=None, **_):
    """CHIRAL_SELECTIVITY dtype is string (VARCHAR-equivalent)."""
    return _check_column_is_string(df, "CHIRAL_SELECTIVITY")


def check_chiral_selectivity_no_whitespace(file_path, df=None, **_):
    """CHIRAL_SELECTIVITY values have no leading/trailing whitespace."""
    return _check_column_no_whitespace(df, "CHIRAL_SELECTIVITY")


def check_chiral_selectivity_no_nulls(file_path, df=None, **_):
    """CHIRAL_SELECTIVITY has no null values."""
    return _check_column_no_nulls(df, "CHIRAL_SELECTIVITY")


def check_chiral_selectivity_in_allowed(file_path, df=None, **_):
    """CHIRAL_SELECTIVITY values are in the allowed list."""
    return _check_column_in_set(df, "CHIRAL_SELECTIVITY", CHIRAL_SELECTIVITY_ALLOWED)


# ---------- MZ checks ----------

# Mass-to-charge ratio range, inclusive on both ends.
MZ_MIN = 150
MZ_MAX = 600


def check_mz_is_float(file_path, df=None, **_):
    """MZ dtype is numeric (FLOAT or int)."""
    return _check_column_is_numeric(df, "MZ")


def check_mz_in_range(file_path, df=None, **_):
    """MZ values are in the valid mass-to-charge range [MZ_MIN, MZ_MAX]."""
    return _check_column_in_range(df, "MZ", MZ_MIN, MZ_MAX)


def check_mz_no_nulls(file_path, df=None, **_):
    """MZ has no null values."""
    return _check_column_no_nulls(df, "MZ")


# ---------- RT checks ----------

# Retention time range (in minutes), exclusive on both ends — spec says
# "more than 0 and less than 6 min".
RT_MIN = 0
RT_MAX = 6


def check_rt_is_float(file_path, df=None, **_):
    """RT dtype is numeric (FLOAT or int)."""
    return _check_column_is_numeric(df, "RT")


def check_rt_in_range(file_path, df=None, **_):
    """RT values are strictly between RT_MIN and RT_MAX (exclusive on both ends)."""
    return _check_column_in_range(
        df, "RT", RT_MIN, RT_MAX,
        lo_inclusive=False, hi_inclusive=False,
    )


def check_rt_no_nulls(file_path, df=None, **_):
    """RT has no null values."""
    return _check_column_no_nulls(df, "RT")


# ---------- LIBRARY_NAME checks ----------

# LIBRARY_NAME must be a single token without underscores or spaces, e.g. "EASMS12kV1".
LIBRARY_NAME_RE = re.compile(r"^[A-Za-z0-9]+$")


def check_library_name_is_string(file_path, df=None, **_):
    """LIBRARY_NAME dtype is string (VARCHAR-equivalent)."""
    return _check_column_is_string(df, "LIBRARY_NAME")


def check_library_name_no_whitespace(file_path, df=None, **_):
    """LIBRARY_NAME values have no leading/trailing whitespace."""
    return _check_column_no_whitespace(df, "LIBRARY_NAME")


def check_library_name_no_nulls(file_path, df=None, **_):
    """LIBRARY_NAME has no null values."""
    return _check_column_no_nulls(df, "LIBRARY_NAME")


def check_library_name_format(file_path, df=None, **_):
    """LIBRARY_NAME is alphanumeric — no underscores or spaces (e.g. 'EASMS12kV1')."""
    ok, fail = _ensure_df_and_columns(df, "LIBRARY_NAME")
    if not ok:
        return False, fail
    invalid = []
    for val in df["LIBRARY_NAME"].dropna().unique():
        if not LIBRARY_NAME_RE.match(str(val).strip()):
            invalid.append(str(val))
    if not invalid:
        return True, "all LIBRARY_NAME values are alphanumeric (no underscores/spaces)"
    preview = invalid[:5]
    suffix = f" (and {len(invalid) - 5} more)" if len(invalid) > 5 else ""
    return False, f"LIBRARY_NAME value(s) with disallowed characters: {preview}{suffix}"


def check_library_name_registered(file_path, df=None, libraries=None, **_):
    """LIBRARY_NAME values match the filename stem of a file in MasterLists/."""
    ok, fail = _ensure_df_and_columns(df, "LIBRARY_NAME")
    if not ok:
        return False, fail
    if libraries is None:
        return False, "libraries list unavailable (MasterLists/ not found)"
    if not libraries:
        return False, "no registered libraries (MasterLists/ is empty)"
    unregistered = []
    for val in df["LIBRARY_NAME"].dropna().unique():
        if str(val).strip() not in libraries:
            unregistered.append(str(val))
    if not unregistered:
        return True, f"all LIBRARY_NAME values are registered ({len(libraries)} libraries available)"
    preview = unregistered[:5]
    suffix = f" (and {len(unregistered) - 5} more)" if len(unregistered) > 5 else ""
    return False, f"unregistered LIBRARY_NAME value(s): {preview}{suffix}"


def check_library_name_consistent(file_path, df=None, **_):
    """All rows share the same LIBRARY_NAME (one file = one library)."""
    ok, fail = _ensure_df_and_columns(df, "LIBRARY_NAME")
    if not ok:
        return False, fail
    non_null = df["LIBRARY_NAME"].dropna()
    if non_null.empty:
        return False, "LIBRARY_NAME column has no non-null values"
    unique = non_null.unique()
    if len(unique) == 1:
        return True, f"all rows share LIBRARY_NAME = '{unique[0]}'"
    preview = list(unique[:5])
    suffix = f" (and {len(unique) - 5} more)" if len(unique) > 5 else ""
    return False, (
        f"LIBRARY_NAME is not consistent: {len(unique)} distinct values: "
        f"{preview}{suffix}"
    )


# ---------- DATA_GENERATOR_NAME checks ----------

def check_data_generator_name_is_string(file_path, df=None, **_):
    """DATA_GENERATOR_NAME dtype is string (VARCHAR-equivalent)."""
    return _check_column_is_string(df, "DATA_GENERATOR_NAME")


def check_data_generator_name_no_whitespace(file_path, df=None, **_):
    """DATA_GENERATOR_NAME values have no leading/trailing whitespace."""
    return _check_column_no_whitespace(df, "DATA_GENERATOR_NAME")


def check_data_generator_name_no_nulls(file_path, df=None, **_):
    """DATA_GENERATOR_NAME has no null values."""
    return _check_column_no_nulls(df, "DATA_GENERATOR_NAME")


def check_data_generator_name_registered(file_path, df=None, data_generators=None, **_):
    """DATA_GENERATOR_NAME matches one of the registered names in Providers.csv
    (column `data_generator_name`)."""
    ok, fail = _ensure_df_and_columns(df, "DATA_GENERATOR_NAME")
    if not ok:
        return False, fail
    if data_generators is None:
        return False, "data-generators list unavailable (Providers.csv not found)"
    if not data_generators:
        return False, (
            "data-generators list is empty (add a `data_generator_name` column "
            "with values to Providers.csv)"
        )
    unregistered = []
    for val in df["DATA_GENERATOR_NAME"].dropna().unique():
        if str(val).strip() not in data_generators:
            unregistered.append(str(val))
    if not unregistered:
        return True, "all DATA_GENERATOR_NAME values are registered"
    preview = unregistered[:5]
    suffix = f" (and {len(unregistered) - 5} more)" if len(unregistered) > 5 else ""
    return False, f"unregistered DATA_GENERATOR_NAME value(s): {preview}{suffix}"


def check_data_generator_name_consistent(file_path, df=None, **_):
    """All rows share the same DATA_GENERATOR_NAME (one file = one generator)."""
    ok, fail = _ensure_df_and_columns(df, "DATA_GENERATOR_NAME")
    if not ok:
        return False, fail
    non_null = df["DATA_GENERATOR_NAME"].dropna()
    if non_null.empty:
        return False, "DATA_GENERATOR_NAME column has no non-null values"
    unique = non_null.unique()
    if len(unique) == 1:
        return True, f"all rows share DATA_GENERATOR_NAME = '{unique[0]}'"
    preview = list(unique[:5])
    suffix = f" (and {len(unique) - 5} more)" if len(unique) > 5 else ""
    return False, (
        f"DATA_GENERATOR_NAME is not consistent: {len(unique)} distinct values: "
        f"{preview}{suffix}"
    )


# ---------- EXPERIMENT_DATE checks ----------

def check_experiment_date_is_string(file_path, df=None, **_):
    """EXPERIMENT_DATE dtype is string (VARCHAR-equivalent)."""
    return _check_column_is_string(df, "EXPERIMENT_DATE")


def check_experiment_date_no_whitespace(file_path, df=None, **_):
    """EXPERIMENT_DATE values have no leading/trailing whitespace."""
    return _check_column_no_whitespace(df, "EXPERIMENT_DATE")


def check_experiment_date_no_nulls(file_path, df=None, **_):
    """EXPERIMENT_DATE has no null values."""
    return _check_column_no_nulls(df, "EXPERIMENT_DATE")


def check_experiment_date_valid_and_not_future(file_path, df=None, **_):
    """EXPERIMENT_DATE values are valid YYYYMMDD strings and not in the future."""
    ok, fail = _ensure_df_and_columns(df, "EXPERIMENT_DATE")
    if not ok:
        return False, fail
    today = datetime.now().date()
    invalid = []   # list of (value, reason)
    for val in df["EXPERIMENT_DATE"].dropna().unique():
        s = str(val).strip()
        try:
            d = datetime.strptime(s, "%Y%m%d").date()
        except ValueError as e:
            invalid.append((s, f"not valid YYYYMMDD: {e}"))
            continue
        if d > today:
            invalid.append((s, f"is in the future (today is {today.isoformat()})"))
    if not invalid:
        return True, "all EXPERIMENT_DATE values are valid YYYYMMDD and not in the future"
    preview = "; ".join(f"'{v}' ({r})" for v, r in invalid[:5])
    suffix = f" (and {len(invalid) - 5} more)" if len(invalid) > 5 else ""
    return False, f"invalid EXPERIMENT_DATE value(s): {preview}{suffix}"


def check_ms_reproducability_is_bool(file_path, df=None, **_):
    """MS_REPRODUCABILITY dtype is boolean."""
    return _check_column_is_bool(df, "MS_REPRODUCABILITY")


def check_ms_reproducability_only_true_false(file_path, df=None, **_):
    """MS_REPRODUCABILITY only contains True or False."""
    return _check_column_in_set(df, "MS_REPRODUCABILITY", {True, False})


def check_ms_reproducability_no_nulls(file_path, df=None, **_):
    """MS_REPRODUCABILITY has no null values."""
    return _check_column_no_nulls(df, "MS_REPRODUCABILITY")


# ---------- POS_INT_REP1 / REP2 / REP3 checks ----------

# Minimum peak intensity expected for a replicate. Set to 0 to allow any
# non-negative intensity; bump if the spec tightens.
POS_INT_REP_MIN = 0


def check_pos_int_rep1_is_float(file_path, df=None, **_):
    """POS_INT_REP1 dtype is numeric (FLOAT or int)."""
    return _check_column_is_numeric(df, "POS_INT_REP1")


def check_pos_int_rep1_at_least(file_path, df=None, **_):
    """POS_INT_REP1 values are >= POS_INT_REP_MIN."""
    return _check_column_at_least(df, "POS_INT_REP1", POS_INT_REP_MIN)


def check_pos_int_rep1_no_nulls(file_path, df=None, **_):
    """POS_INT_REP1 has no null values."""
    return _check_column_no_nulls(df, "POS_INT_REP1")


def check_pos_int_rep2_is_float(file_path, df=None, **_):
    """POS_INT_REP2 dtype is numeric (FLOAT or int)."""
    return _check_column_is_numeric(df, "POS_INT_REP2")


def check_pos_int_rep2_at_least(file_path, df=None, **_):
    """POS_INT_REP2 values are >= POS_INT_REP_MIN."""
    return _check_column_at_least(df, "POS_INT_REP2", POS_INT_REP_MIN)


def check_pos_int_rep2_no_nulls(file_path, df=None, **_):
    """POS_INT_REP2 has no null values."""
    return _check_column_no_nulls(df, "POS_INT_REP2")


def check_pos_int_rep3_is_float(file_path, df=None, **_):
    """POS_INT_REP3 dtype is numeric (FLOAT or int)."""
    return _check_column_is_numeric(df, "POS_INT_REP3")


def check_pos_int_rep3_at_least(file_path, df=None, **_):
    """POS_INT_REP3 values are >= POS_INT_REP_MIN."""
    return _check_column_at_least(df, "POS_INT_REP3", POS_INT_REP_MIN)


def check_pos_int_rep3_no_nulls(file_path, df=None, **_):
    """POS_INT_REP3 has no null values."""
    return _check_column_no_nulls(df, "POS_INT_REP3")


# ---------- BINARY_LABEL checks ----------

def check_binary_label_is_int(file_path, df=None, **_):
    """BINARY_LABEL dtype is integer."""
    return _check_column_is_int(df, "BINARY_LABEL")


def check_binary_label_only_zero_one(file_path, df=None, **_):
    """BINARY_LABEL contains only 0 or 1."""
    return _check_column_in_set(df, "BINARY_LABEL", {0, 1})


def check_binary_label_no_nulls(file_path, df=None, **_):
    """BINARY_LABEL has no null values."""
    return _check_column_no_nulls(df, "BINARY_LABEL")


def check_protein_conc_is_float(file_path, df=None, **_):
    """PROTEIN_CONC dtype is numeric (FLOAT or int)."""
    return _check_column_is_numeric(df, "PROTEIN_CONC")


def check_protein_conc_equals_expected(file_path, df=None, **_):
    """PROTEIN_CONC values equal the expected concentration (PROTEIN_CONC_EXPECTED)."""
    return _check_column_equals(df, "PROTEIN_CONC", PROTEIN_CONC_EXPECTED)


def check_protein_conc_no_nulls(file_path, df=None, **_):
    """PROTEIN_CONC has no null values."""
    return _check_column_no_nulls(df, "PROTEIN_CONC")


def check_incubation_volume_is_float(file_path, df=None, **_):
    """INCUBATION_VOLUME dtype is numeric (FLOAT or int)."""
    return _check_column_is_numeric(df, "INCUBATION_VOLUME")


def check_incubation_volume_positive(file_path, df=None, **_):
    """INCUBATION_VOLUME values are strictly positive (> 0)."""
    return _check_column_positive(df, "INCUBATION_VOLUME")


def check_incubation_volume_no_nulls(file_path, df=None, **_):
    """INCUBATION_VOLUME has no null values."""
    return _check_column_no_nulls(df, "INCUBATION_VOLUME")


def check_protein_id_consistent_per_target(file_path, df=None, **_):
    """Each TARGET_ID maps to exactly one PROTEIN_ID across all its rows."""
    ok, fail = _ensure_df_and_columns(df, "TARGET_ID", "PROTEIN_ID")
    if not ok:
        return False, fail
    distinct_per_target = df.groupby("TARGET_ID")["PROTEIN_ID"].nunique()
    inconsistent = distinct_per_target[distinct_per_target > 1]
    if inconsistent.empty:
        return True, (
            f"each TARGET_ID maps to a single PROTEIN_ID "
            f"({len(distinct_per_target)} targets checked)"
        )
    preview_lines = []
    for tid in inconsistent.index[:5]:
        ids = df.loc[df["TARGET_ID"] == tid, "PROTEIN_ID"].dropna().unique().tolist()
        preview_lines.append(f"'{tid}' -> {ids}")
    suffix = (
        f" (and {len(inconsistent) - 5} more)"
        if len(inconsistent) > 5 else ""
    )
    return False, (
        f"{len(inconsistent)} TARGET_ID(s) map to multiple PROTEIN_IDs: "
        f"{'; '.join(preview_lines)}{suffix}"
    )


# ---------- POOL_SIZE checks ----------

POOL_SIZE_MIN = 400
POOL_SIZE_MAX = 1500


def check_pool_size_is_int(file_path, df=None, **_):
    """POOL_SIZE dtype is integer (or a float column whose values are all whole numbers)."""
    return _check_column_is_int(df, "POOL_SIZE")


def check_pool_size_in_range(file_path, df=None, **_):
    """POOL_SIZE values are positive integers in [POOL_SIZE_MIN, POOL_SIZE_MAX]."""
    return _check_column_in_range(df, "POOL_SIZE", POOL_SIZE_MIN, POOL_SIZE_MAX)


def check_pool_size_no_nulls(file_path, df=None, **_):
    """POOL_SIZE has no null values."""
    return _check_column_no_nulls(df, "POOL_SIZE")


def check_asms_batch_name_format(file_path, df=None, providers=None, **_):
    """ASMS_BATCH_NAME values follow '<provider>_<batch_number>' with a
    provider acronym registered in Providers.csv."""
    ok, fail = _ensure_df_and_columns(df, "ASMS_BATCH_NAME")
    if not ok:
        return False, fail
    if providers is None:
        return False, "providers list unavailable (Providers.csv not found)"
    if not providers:
        return False, "providers list is empty"

    invalid = []   # list of (value, reason)
    for val in df["ASMS_BATCH_NAME"].dropna().unique():
        s = str(val).strip()
        m = _BATCH_NAME_RE.match(s)
        if not m:
            invalid.append((s, "does not match '<provider>_<batch_number>'"))
            continue
        provider = m.group(1).lower()
        if provider not in providers:
            invalid.append((s, f"provider '{provider}' is not registered"))

    if not invalid:
        return True, (
            "all ASMS_BATCH_NAME values follow '<provider>_<batch_number>' "
            "with a registered provider"
        )

    preview = "; ".join(f"'{v}' ({reason})" for v, reason in invalid[:5])
    suffix = f" (and {len(invalid) - 5} more)" if len(invalid) > 5 else ""
    return False, f"invalid ASMS_BATCH_NAME value(s): {preview}{suffix}"


# ---------- Section registry ----------

SECTIONS = [
    ("File Format Checks", [
        ("File opens without errors",                check_file_opens),
        ("File is a CSV (extension)",                check_is_csv),
        ("File is not empty",                        check_file_not_empty),
        (f"File size is under {MAX_FILE_SIZE_GB} GB", check_file_size_under_limit),
        ("File encoding is UTF-8",                   check_encoding_is_utf8),
        ("File is a CSV (parseable content)",        check_csv_parseable),
        ("Columns match ASMS Meta Data.csv reference", check_columns_match_metadata),
    ]),
    ("Filename Format Checks", [
        ("Filename has no special characters or spaces", check_filename_no_special_chars),
        ("Filename starts with 'asms_'",                 check_filename_starts_with_asms),
        ("Filename matches overall format",              check_filename_overall_format),
        ("Provider acronym is registered",               check_provider_acronym),
        ("Batch number is in valid range",               check_batch_number_range),
        ("Library name is registered",                   check_library_name),
        ("Date is valid YYYYMMDD and not in the future", check_date_valid_and_not_future),
    ]),
    ("Row Content Checks", [
        ("No fully duplicate rows", check_no_duplicate_rows),
    ]),
    ("Column Content Checks", [
        # COMPOUND_ID
        ("COMPOUND_ID is string (VARCHAR)",                  check_compound_id_is_string),
        ("COMPOUND_ID has no leading/trailing whitespace",   check_compound_id_no_whitespace),
        ("COMPOUND_ID has no null values",                   check_compound_id_no_nulls),
        ("COMPOUND_ID is unique within each TARGET_ID",      check_compound_id_unique_per_target),
        # SMILES
        ("SMILES is string (VARCHAR)",                       check_smiles_is_string),
        ("SMILES has no leading/trailing whitespace",        check_smiles_no_whitespace),
        ("SMILES has no null values",                        check_smiles_no_nulls),
        ("SMILES is valid (non-empty, RDKit-parseable)",     check_smiles_valid),
        ("SMILES is in the associated library",              check_smiles_in_library),
        ("SMILES is unique within each TARGET_ID",           check_smiles_unique_per_target),
        # ASMS_BATCH_NAME
        ("ASMS_BATCH_NAME is string (VARCHAR)",                check_asms_batch_name_is_string),
        ("ASMS_BATCH_NAME has no leading/trailing whitespace", check_asms_batch_name_no_whitespace),
        ("ASMS_BATCH_NAME has no null values",                 check_asms_batch_name_no_nulls),
        ("ASMS_BATCH_NAME is consistent across all rows",      check_asms_batch_name_consistent),
        ("ASMS_BATCH_NAME follows <provider>_<batch_number>",  check_asms_batch_name_format),
        # COMPOUND_FORMULA
        ("COMPOUND_FORMULA is string (VARCHAR)",                 check_compound_formula_is_string),
        ("COMPOUND_FORMULA has no leading/trailing whitespace",  check_compound_formula_no_whitespace),
        ("COMPOUND_FORMULA has no null values",                  check_compound_formula_no_nulls),
        ("COMPOUND_FORMULA is in the associated library",        check_compound_formula_in_library),
        # POOL_NAME
        ("POOL_NAME is string (VARCHAR)",                        check_pool_name_is_string),
        ("POOL_NAME has no leading/trailing whitespace",         check_pool_name_no_whitespace),
        ("POOL_NAME has no null values",                         check_pool_name_no_nulls),
        # POOL_ID
        ("POOL_ID is string (VARCHAR)",                          check_pool_id_is_string),
        ("POOL_ID has no leading/trailing whitespace",           check_pool_id_no_whitespace),
        ("POOL_ID has no null values",                           check_pool_id_no_nulls),
        # POOL_SIZE
        ("POOL_SIZE is integer (INT)",                                                  check_pool_size_is_int),
        (f"POOL_SIZE is in valid range [{POOL_SIZE_MIN}, {POOL_SIZE_MAX}]",             check_pool_size_in_range),
        ("POOL_SIZE has no null values",                                                check_pool_size_no_nulls),
        # TARGET_ID
        ("TARGET_ID is string (VARCHAR)",                                check_target_id_is_string),
        ("TARGET_ID has no leading/trailing whitespace",                 check_target_id_no_whitespace),
        ("TARGET_ID has no null values",                                 check_target_id_no_nulls),
        ("TARGET_ID matches <name>_<UniprotID>_<start>_<end>",           check_target_id_format),
        ("TARGET_ID start < end (and both numeric)",                     check_target_id_start_lt_end),
        ("All TARGET_IDs have the same number of compounds",             check_target_id_same_compound_count),
        # PROTEIN_ID
        ("PROTEIN_ID is string (VARCHAR)",                               check_protein_id_is_string),
        ("PROTEIN_ID has no leading/trailing whitespace",                check_protein_id_no_whitespace),
        ("PROTEIN_ID has no null values",                                check_protein_id_no_nulls),
        ("PROTEIN_ID is consistent within each TARGET_ID",               check_protein_id_consistent_per_target),
        # INCUBATION_VOLUME
        ("INCUBATION_VOLUME is numeric (FLOAT)",                         check_incubation_volume_is_float),
        ("INCUBATION_VOLUME values are positive (> 0)",                  check_incubation_volume_positive),
        ("INCUBATION_VOLUME has no null values",                         check_incubation_volume_no_nulls),
        # PROTEIN_CONC
        ("PROTEIN_CONC is numeric (FLOAT)",                              check_protein_conc_is_float),
        (f"PROTEIN_CONC equals expected value ({PROTEIN_CONC_EXPECTED})", check_protein_conc_equals_expected),
        ("PROTEIN_CONC has no null values",                              check_protein_conc_no_nulls),
        # MS_REPRODUCABILITY
        ("MS_REPRODUCABILITY is boolean (BOOL)",                         check_ms_reproducability_is_bool),
        ("MS_REPRODUCABILITY only contains True/False",                  check_ms_reproducability_only_true_false),
        ("MS_REPRODUCABILITY has no null values",                        check_ms_reproducability_no_nulls),
        # POS_INT_REP1
        ("POS_INT_REP1 is numeric (FLOAT)",                              check_pos_int_rep1_is_float),
        (f"POS_INT_REP1 values are >= {POS_INT_REP_MIN}",                check_pos_int_rep1_at_least),
        ("POS_INT_REP1 has no null values",                              check_pos_int_rep1_no_nulls),
        # POS_INT_REP2
        ("POS_INT_REP2 is numeric (FLOAT)",                              check_pos_int_rep2_is_float),
        (f"POS_INT_REP2 values are >= {POS_INT_REP_MIN}",                check_pos_int_rep2_at_least),
        ("POS_INT_REP2 has no null values",                              check_pos_int_rep2_no_nulls),
        # POS_INT_REP3
        ("POS_INT_REP3 is numeric (FLOAT)",                              check_pos_int_rep3_is_float),
        (f"POS_INT_REP3 values are >= {POS_INT_REP_MIN}",                check_pos_int_rep3_at_least),
        ("POS_INT_REP3 has no null values",                              check_pos_int_rep3_no_nulls),
        # BINARY_LABEL
        ("BINARY_LABEL is integer (INT)",                                check_binary_label_is_int),
        ("BINARY_LABEL only contains {0, 1}",                            check_binary_label_only_zero_one),
        ("BINARY_LABEL has no null values",                              check_binary_label_no_nulls),
        # LIBRARY_NAME
        ("LIBRARY_NAME is string (VARCHAR)",                             check_library_name_is_string),
        ("LIBRARY_NAME has no leading/trailing whitespace",              check_library_name_no_whitespace),
        ("LIBRARY_NAME has no null values",                              check_library_name_no_nulls),
        ("LIBRARY_NAME is alphanumeric (no underscores/spaces)",         check_library_name_format),
        ("LIBRARY_NAME is registered",                                   check_library_name_registered),
        ("LIBRARY_NAME is consistent across all rows",                   check_library_name_consistent),
        # DATA_GENERATOR_NAME
        ("DATA_GENERATOR_NAME is string (VARCHAR)",                      check_data_generator_name_is_string),
        ("DATA_GENERATOR_NAME has no leading/trailing whitespace",       check_data_generator_name_no_whitespace),
        ("DATA_GENERATOR_NAME has no null values",                       check_data_generator_name_no_nulls),
        ("DATA_GENERATOR_NAME is registered (in Providers.csv)",         check_data_generator_name_registered),
        ("DATA_GENERATOR_NAME is consistent across all rows",            check_data_generator_name_consistent),
        # EXPERIMENT_DATE
        ("EXPERIMENT_DATE is string (VARCHAR)",                          check_experiment_date_is_string),
        ("EXPERIMENT_DATE has no leading/trailing whitespace",           check_experiment_date_no_whitespace),
        ("EXPERIMENT_DATE has no null values",                           check_experiment_date_no_nulls),
        ("EXPERIMENT_DATE is valid YYYYMMDD and not in the future",      check_experiment_date_valid_and_not_future),
        # CHIRAL_SELECTIVITY
        ("CHIRAL_SELECTIVITY is string (VARCHAR)",                       check_chiral_selectivity_is_string),
        ("CHIRAL_SELECTIVITY has no leading/trailing whitespace",        check_chiral_selectivity_no_whitespace),
        ("CHIRAL_SELECTIVITY has no null values",                        check_chiral_selectivity_no_nulls),
        ("CHIRAL_SELECTIVITY is one of the allowed values",              check_chiral_selectivity_in_allowed),
        # MZ
        ("MZ is numeric (FLOAT)",                                        check_mz_is_float),
        (f"MZ is in valid range [{MZ_MIN}, {MZ_MAX}]",                   check_mz_in_range),
        ("MZ has no null values",                                        check_mz_no_nulls),
        # RT
        ("RT is numeric (FLOAT)",                                        check_rt_is_float),
        (f"RT is in valid range ({RT_MIN}, {RT_MAX}) exclusive",         check_rt_in_range),
        ("RT has no null values",                                        check_rt_no_nulls),
        # PROTEIN_SEQ
        ("PROTEIN_SEQ is string (VARCHAR)",                              check_protein_seq_is_string),
        ("PROTEIN_SEQ has no leading/trailing whitespace",               check_protein_seq_no_whitespace),
        ("PROTEIN_SEQ has no null values",                               check_protein_seq_no_nulls),
        (f"PROTEIN_SEQ length > {PROTEIN_SEQ_MIN_LENGTH}",                check_protein_seq_min_length),
        # PROTEIN_TAG
        ("PROTEIN_TAG is string (VARCHAR)",                              check_protein_tag_is_string),
        ("PROTEIN_TAG has no leading/trailing whitespace",               check_protein_tag_no_whitespace),
        ("PROTEIN_TAG has no null values",                               check_protein_tag_no_nulls),
    ]),
]


# ---------- Orchestrator ----------

SEPARATOR = "=" * 60


def run_quality_checks(file_path, log_dir, providers_csv=None, masterlist_dir=None,
                       meta_csv=None):
    """
    Run every check in SECTIONS against `file_path` and write a sectioned log.

    Args:
        file_path (str):       Path to the raw input file being validated.
        log_dir (str):         Directory where the log file is written.
        providers_csv (str):   Path to Providers.csv (acronym list). Optional.
        masterlist_dir (str):  Path to the MasterLists/ folder. Optional.
        meta_csv (str):        Path to ASMS Meta Data.csv (valid columns). Optional.

    Returns:
        tuple[bool, list[tuple[str, str]]]:
            (all_passed, failed_checks). `failed_checks` is a list of
            (description, message) pairs for every check that failed; empty
            when all_passed is True.
    """
    os.makedirs(log_dir, exist_ok=True)
    file_name = os.path.basename(file_path)
    base_name = os.path.splitext(file_name)[0]
    today = datetime.now().strftime("%Y%m%d")
    log_path = os.path.join(log_dir, f"QCaircheck{today}_{base_name}.log")

    # Pre-load context that several checks need. `output_dir` lets checks
    # write supplementary files (e.g. FullyDuplicate_rows_report.csv) next to the log.
    # `df` is the input CSV loaded once with fully-duplicate rows dropped, so
    # column-content checks operate on the same cleaned data the pipeline would
    # see after Step 3's drop_duplicates().
    context = {
        "providers":       _load_providers(providers_csv),
        "data_generators": _load_data_generators(providers_csv),
        "libraries":       _list_libraries(masterlist_dir),
        "meta_columns":    _load_meta_columns(meta_csv),
        "output_dir":      log_dir,
        "df":              _load_dataframe(file_path),
        # Raw path passed through so library-aware checks can look up
        # MasterList_Information.xlsx and load the relevant library file.
        "masterlist_dir":  masterlist_dir,
    }

    all_passed = True
    failed_checks = []   # collected for the caller to print a nice console message
    rows = []            # accumulated for the Excel report
    check_idx = 0
    generated_at = datetime.now().isoformat(timespec="seconds")
    with open(log_path, "w", encoding="utf-8") as log:
        log.write("Quality Check Log\n")
        log.write(f"File:      {file_name}\n")
        log.write(f"Generated: {generated_at}\n")
        log.write(SEPARATOR + "\n")

        for section_name, checks in SECTIONS:
            log.write("\n")
            log.write(f"{section_name}\n")
            log.write(SEPARATOR + "\n\n")

            for description, check_fn in checks:
                check_idx += 1
                try:
                    result = check_fn(file_path, **context)
                    # Checks can return 2-tuple (passed, msg) or 3-tuple
                    # (passed, msg, status) where status is "PASS", "FAIL", or "WARN".
                    if isinstance(result, tuple) and len(result) == 3:
                        passed, message, status = result
                    else:
                        passed, message = result
                        status = "PASS" if passed else "FAIL"
                except Exception as e:
                    passed, message, status = False, f"check raised an exception: {e}", "FAIL"
                log.write(f"Check {check_idx}: {description}\n")
                log.write(f"  Result : {status}\n")
                log.write(f"  Detail : {message}\n\n")
                rows.append({
                    "Section":  section_name,
                    "Check #":  check_idx,
                    "Criteria": description,
                    "Status":   status,
                    "Detail":   message,
                })
                if status == "FAIL":
                    all_passed = False
                    failed_checks.append((description, message))

        log.write(SEPARATOR + "\n")
        log.write(f"Overall : {'PASS' if all_passed else 'FAIL'}\n")

    # Write the Excel companion next to the .log
    excel_path = os.path.join(log_dir, f"QCaircheck{today}_{base_name}.xlsx")
    try:
        _write_excel_report(
            rows=rows,
            excel_path=excel_path,
            file_name=file_name,
            generated_at=generated_at,
            overall_status="PASS" if all_passed else "FAIL",
        )
    except Exception as e:
        # Excel write failure should not block QC; record in log
        with open(log_path, "a", encoding="utf-8") as log:
            log.write(f"\n(Note: failed to write Excel report: {e})\n")

    return all_passed, failed_checks


def _write_excel_report(rows, excel_path, file_name, generated_at, overall_status):
    """Write a color-coded Excel version of the QC log.

    Layout (one sheet, "QC Results"):
      Row 1: File:      <file_name>
      Row 2: Generated: <iso timestamp>
      Row 3: Overall:   <PASS/FAIL>
      Row 4: (blank)
      Row 5: Header (Section / Check # / Criteria / Status / Detail)
      Row 6+: Data, with each row's background color tied to its Status.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "QC Results"

    # Metadata block
    bold = Font(bold=True)
    ws["A1"] = "File:";      ws["B1"] = file_name
    ws["A2"] = "Generated:"; ws["B2"] = generated_at
    ws["A3"] = "Overall:";   ws["B3"] = overall_status
    for addr in ("A1", "A2", "A3"):
        ws[addr].font = bold

    # Header row
    header_row = 5
    headers = ["Section", "Check #", "Criteria", "Status", "Detail"]
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=name)
        cell.font = bold

    # Data rows with color coding by status
    status_fills = {
        "PASS": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "FAIL": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "WARN": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    }
    wrap = Alignment(wrap_text=True, vertical="top")
    for offset, row in enumerate(rows, start=1):
        excel_row = header_row + offset
        values = [row["Section"], row["Check #"], row["Criteria"], row["Status"], row["Detail"]]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.fill = status_fills.get(row["Status"], status_fills["PASS"])
            cell.alignment = wrap

    # Column widths
    widths = {"A": 26, "B": 9, "C": 50, "D": 8, "E": 100}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Freeze the header row so it stays visible when scrolling
    ws.freeze_panes = f"A{header_row + 1}"

    wb.save(excel_path)
