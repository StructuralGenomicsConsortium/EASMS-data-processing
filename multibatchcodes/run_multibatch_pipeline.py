# -*- coding: utf-8 -*-
"""One-shot multi-batch pipeline: raw per-protein CSVs -> all analysis columns + plots.

This replaces the chain of one-off scripts (compute_enrichment_median.py ->
compute_enrichment_median_normalized.py -> compute_enrichment_percentile.py ->
compute_bead_ratio.py -> compute_modified5_20to36.py -> hit_matrix_* ->
export_matrix_*) with a single entry point. When a new set of batches arrives,
point it at the folder and everything is produced in one run.

    python multibatchcodes/run_multibatch_pipeline.py --input-dir MultiBatch_20to40

Folders (all derived from --input-dir, override with the flags):

    MultiBatch_20to40/            input  - one CSV per protein (the raw columns)
    MultiBatch_20to40_modified/   output - same files + every computed column
    MultiBatch_20to40_results/    output - heatmap PNG + full matrix CSV/XLSX + cache

Columns added to every row (35 of them):

    fold_articlebased_within                 article eq.1, within batch (see below)
    target_median, signal_flag
    raw median   within/across : nontarget_median_*, enrichment_*, pvalue_*, label_*
    median-norm  (_norm)       : + norm_scale_factor
    percentile-norm (_pctnorm)
    bead         : bead_signal, bead_detected, bead_ratio, bead_binder_flag_r5/r10
    Normalized score1 + ns1_scale_factor / ns1_target_median /
                        ns1_enrichment_within / ns1_log2_enrichment

"Normalized score1" is its own chain and does NOT reuse the median-norm columns:
its batch scaling divides by the median OF THE BATCH MEDIANS (the median-norm
family divides by the pooled median of all detected values), and it scales the
3000s along with everything else. See normalized_score1() for the full order.

Backgrounds: "within" = the other proteins of the SAME `ASMS_BATCH_NAME`;
"across" = every other protein in the folder, all batches pooled.

One deliberate speed shortcut, mathematically exact: median-norm is a per-batch
LINEAR (strictly increasing) scale, so within-batch ratios AND ranks are
unchanged -> the *_within_norm columns are derived from the raw ones.

No such shortcut for percentile-norm. It is monotonic but NOT strictly
increasing: the pooled reference has a large plateau at the 3000 floor, so
distinct raw intensities collapse onto equal transformed values. Those new ties
change Mann-Whitney ranks, so pvalue_within_pctnorm must be computed on the
transformed values (on 20to22 the shortcut moved 15.5% of rows and flipped 34
hit calls; compute_modified5_20to36.py still has it).

Across-batch always mixes batches, so both norms recompute it in full.

MISSING REPLICATES -- two conventions live side by side, on purpose:
  * The median/norm/pctnorm/bead family imputes a missing replicate as the 3000
    FLOOR ("not detected"). Without it a NaN propagates through the percentile
    transform and the across-protein median and wipes out whole columns.
  * `fold_articlebased_within` instead EXCLUDES an unmeasured protein from its
    batch background (so N shrinks) and is NaN for the unmeasured row itself.
    That matches the pipeline's own `fold_change`, which is NaN on those rows.
  Both are documented in compute_fold_articlebased_within.py; on 20to36 the
  choice moves 67 rows by a few percent and flips no hit call.

Plots/exports (in the results folder), reproducing the previous scripts:
  <name>_hit_matrix_intensity_sorted_floor.png   molecules x proteins heatmap
  <name>_matrix_sorted_floor_full.csv / .xlsx    the same matrix with all columns

Useful flags:
    --write-only     re-write outputs from the cached columns (no ~35 min recompute)
    --skip-columns   leave the _modified folder alone, only rebuild plots/matrix
    --skip-heatmap / --skip-matrix
    --beads PATH     bead reference workbook (default MultiBatchResults/beads_clean.xlsx)
    --validate       cross-check the vectorised maths against direct recomputation
"""

import os
import sys
import glob
import argparse
import functools
import warnings
import statistics
from collections import defaultdict

import numpy as np
import pandas as pd
from scipy.stats import mannwhitneyu, rankdata

print = functools.partial(print, flush=True)  # live progress on long runs

REP = ["POS_INT_REP1", "POS_INT_REP2", "POS_INT_REP3"]
KEYS = ["ASMS_BATCH_NAME", "COMPOUND_ID"]
FLOOR = 3000.0
STRONG_MIN = 1e6
MODERATE_MIN = 1e4
ENR_THRESHOLD = 5.0
P_THRESHOLD = 0.05
C5 = ["5Beads_V1P0113_1", "5Beads_V1P0113_2", "5Beads_V1P0113_3"]
C10 = ["10Beads_V1P0113_1", "10Beads_V1P0113_2", "10Beads_V1P0113_3"]

# every column this pipeline appends, in output order
NEW_COLS = [
    "fold_articlebased_within",
    "target_median", "signal_flag",
    "nontarget_median_within", "enrichment_within", "pvalue_within", "label_within",
    "nontarget_median_across", "enrichment_across", "pvalue_across", "label_across",
    "norm_scale_factor", "target_median_norm",
    "nontarget_median_within_norm", "enrichment_within_norm",
    "pvalue_within_norm", "label_within_norm",
    "nontarget_median_across_norm", "enrichment_across_norm",
    "pvalue_across_norm", "label_across_norm",
    "target_median_pctnorm",
    "nontarget_median_within_pctnorm", "enrichment_within_pctnorm",
    "pvalue_within_pctnorm", "label_within_pctnorm",
    "nontarget_median_across_pctnorm", "enrichment_across_pctnorm",
    "pvalue_across_pctnorm", "label_across_pctnorm",
    "bead_signal", "bead_detected", "bead_ratio",
    "bead_binder_flag_r5", "bead_binder_flag_r10",
    "ns1_scale_factor", "ns1_target_median",
    "ns1_nontarget_mean_within", "ns1_enrichment_within",
    "ns1_nontarget_mean_across", "ns1_enrichment_across",
    "ns1_log2_enrichment", "ns1_ref_median", "ns1_ref_mad",
    # ----------------------------------------------------------------------- #
    #  LAST COLUMNS: the across-batch enrichment 2x2, under explicit names,
    #  followed by the score. Two independent choices, so four columns:
    #
    #                     |  median background      |  mean background
    #     ----------------+-------------------------+--------------------------
    #     median scaling  |  ..medscale_medbg       |  ..medscale_meanbg
    #     quantile scaling|  ..quantscale_medbg     |  ..quantscale_meanbg
    #
    #  "scaling"    = how batches are put on a common scale before pooling:
    #                 medscale   -> one factor per batch, reference/batch median
    #                              of the DETECTED (>3000) values
    #                 quantscale -> each batch's ranks mapped onto the pooled
    #                              reference distribution (full quantile norm)
    #  "background" = how the other proteins of the same compound, ALL batches
    #                 pooled, are combined into one denominator: their median or
    #                 their mean.
    #
    #  Three are the same numbers as older columns, renamed for clarity, so
    #  nothing about them changes -- the old names stay because four other
    #  scripts read them:
    #     medscale_medbg      == enrichment_across_norm
    #     medscale_meanbg     == ns1_enrichment_across
    #     quantscale_medbg    == enrichment_across_pctnorm
    #     quantscale_meanbg   == NEW, the cell that was missing
    #  (`enrichment_across` remains the UNSCALED median-background version, kept
    #  out of the 2x2 as the no-correction baseline.)
    # ----------------------------------------------------------------------- #
    "enrich_across_medscale_medbg",
    "enrich_across_medscale_meanbg",
    "enrich_across_quantscale_medbg",
    "enrich_across_quantscale_meanbg",
    "Normalized score1",
]


# --------------------------------------------------------------------------- #
#  article equation 1 -- within-batch enrichment fold
# --------------------------------------------------------------------------- #

def peak_intensity(reps):
    """A = mean of the replicate peak intensities; all-missing -> NaN."""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", RuntimeWarning)   # "Mean of empty slice"
        return np.nanmean(reps, axis=1)


def article_fold_within(big, reps_raw):
    """fold = A_POI * N / sum(A_i), background = same batch, same compound.

    N counts only proteins with a MEASURED intensity, so an unmeasured protein
    neither counts toward N nor contributes to the sum (it would otherwise
    inflate the fold). Uses the RAW replicates -- no floor imputation.
    """
    A = peak_intensity(reps_raw)
    tmp = big[KEYS].copy()
    tmp["_A"] = A
    stats = tmp.groupby(KEYS, sort=False)["_A"].agg(S="sum", K="count").reset_index()
    m = big[KEYS].merge(stats, on=KEYS, how="left")       # left merge keeps row order
    N = m["K"].to_numpy() - 1.0
    denom = m["S"].to_numpy() - A
    with np.errstate(divide="ignore", invalid="ignore"):
        fold = A * N / denom
    return np.where((N <= 0) | (denom <= 0), np.nan, fold)


# --------------------------------------------------------------------------- #
#  median / normalized / percentile enrichment + Mann-Whitney
# --------------------------------------------------------------------------- #

def signal_flags(reps):
    n_strong = (reps >= STRONG_MIN).sum(axis=1)
    n_mod = ((reps >= MODERATE_MIN) & (reps < STRONG_MIN)).sum(axis=1)
    flag = np.full(len(reps), "none", dtype=object)
    m = n_mod > 0
    flag[m] = [f"{k} moderate" for k in n_mod[m]]
    s = n_strong > 0
    flag[s] = [f"{k} strong" for k in n_strong[s]]
    return flag


def enrichment_and_p(reps, tmed, idx_by_comp, batch, ndet,
                     do_within=True, within_pvalue=True,
                     do_across=True, across_pvalue=True):
    """Per-protein-median background enrichment + Mann-Whitney, within & across.

    Flags let callers skip work known to be redundant (see module docstring).
    Returns nt_within, nt_across, p_within, p_across (NaN where not computed).
    """
    n = len(tmed)
    nt_w = np.full(n, np.nan); nt_a = np.full(n, np.nan)
    p_w = np.full(n, np.nan); p_a = np.full(n, np.nan)
    for c, idxs in idx_by_comp.items():
        idxs = np.asarray(idxs)
        vals = tmed[idxs]
        bts = batch[idxs]
        for j, i in enumerate(idxs):
            if do_across:
                others = np.delete(vals, j)
                if others.size:
                    nt_a[i] = np.median(others)
                    if across_pvalue:
                        p_a[i] = 1.0 if ndet[i] == 0 else mannwhitneyu(
                            reps[i], others, alternative="two-sided").pvalue
            if do_within:
                same = (bts == bts[j]); same[j] = False
                within = vals[same]
                if within.size:
                    nt_w[i] = np.median(within)
                    if within_pvalue:
                        p_w[i] = 1.0 if ndet[i] == 0 else mannwhitneyu(
                            reps[i], within, alternative="two-sided").pvalue
    return nt_w, nt_a, p_w, p_a


def labels(enr, pval):
    return ((enr >= ENR_THRESHOLD) & (pval <= P_THRESHOLD)).astype(int)


# --------------------------------------------------------------------------- #
#  "Normalized score1" -- batch scaling -> within-batch enrichment -> robust Z
# --------------------------------------------------------------------------- #

MAD_WELL_FRAC = 0.75    # "well detected" = signal on >= this fraction of a
                        #   compound's proteins
MAD_PCT = 5.0           # percentile of the well-detected MAD distribution to take
MAD_MIN_FALLBACK = 0.12  # used only if the data cannot support an estimate
MAD_EPS = 1e-9          # a MAD at or below this is a floating-point zero
MAD_PLAUSIBLE = (0.02, 0.50)   # outside this band the estimate is only WARNED about


def estimate_mad_min(ref_mad, detected, comp,
                     well_frac=MAD_WELL_FRAC, pct=MAD_PCT,
                     fallback=MAD_MIN_FALLBACK, min_rows=200, verbose=True):
    """Derive the MAD floor FROM THE DATA, so more batches re-derive it.

    Why a floor is needed at all is in normalized_score1's docstring. The value
    must not be hard-coded, because it depends on the panel: with 24 proteins per
    compound 43.8% of reference MADs are degenerate, with 136 only 15.0%, so the
    distribution of genuine spreads is not the same dataset to dataset.

    The procedure, and it is the whole definition of the number:

      1. take the leave-one-out reference MAD of every row (computed already)
      2. keep only WELL-DETECTED compounds -- signal on >= `well_frac` of that
         compound's proteins. This is the crucial step: in the poorly-detected
         strata 78-99.7% of MADs are exactly 0 (whole batches float together, so
         every L collapses to one value), and including them would just measure
         the censoring. In the well-detected stratum that share falls to ~0.1%,
         so the MAD there is genuine biological spread.
      3. drop any residual floating-point zeros
      4. take the `pct`-th percentile

    The result reads as "the smallest spread this dataset itself says is real".
    Anything below it is treated as artifact and cannot shrink the denominator.

    Returns (mad_min, n_rows_used).
    """
    frac_det = (pd.Series(np.asarray(detected, dtype=float))
                .groupby(pd.Series(comp)).transform("mean").to_numpy())
    ok = (frac_det >= well_frac) & np.isfinite(ref_mad) & (ref_mad > MAD_EPS)
    n_ok = int(ok.sum())

    if n_ok < min_rows:
        if verbose:
            print(f"  MAD_MIN: only {n_ok:,} well-detected rows with a non-zero "
                  f"MAD (need {min_rows:,}) -> falling back to {fallback}")
        return float(fallback), n_ok

    mad_min = float(np.percentile(ref_mad[ok], pct))
    if verbose:
        share_zero = float((ref_mad[frac_det >= well_frac] <= MAD_EPS).mean())
        print(f"  MAD_MIN estimated from the data = {mad_min:.4f} log2 units "
              f"(= {2 ** mad_min:.3f}x, i.e. {(2 ** mad_min - 1) * 100:.1f}% spread)")
        print(f"    p{pct:g} of the leave-one-out MAD over {n_ok:,} rows whose "
              f"compound is detected on >={well_frac:.0%} of its proteins")
        print(f"    (degenerate MAD share in that stratum: {share_zero:.1%} -- "
              f"if this is not small the stratum is not clean)")
        lo, hi = MAD_PLAUSIBLE
        if not (lo <= mad_min <= hi):
            print(f"    [WARNING] {mad_min:.4f} is outside the plausible band "
                  f"[{lo}, {hi}]. Not clamped -- inspect ns1_ref_mad before "
                  f"trusting the score.")
    return mad_min, n_ok


def normalized_score1(reps, batch, comp, mad_min=None):
    """The requested chain, in order:

        1. exclude the 3000 floor when estimating each batch median
        2. batch median      = median of the DETECTED replicate values of that batch
        3. reference         = median OF THE BATCH MEDIANS (not the pooled median)
        4. scale factor      = reference / batch median
        5. scale ALL values of the batch, the 3000s included
        6. M_ipb             = median of the 3 scaled replicates
        7. E_ipb             = M_ipb / mean_{q!=p} [ M_iqb ]     (same batch)
        8. L_ipb             = log2(E_ipb)
           S_ip^robust       = ( L_ip - median_{q!=p}(L_iq) )
                               / ( 1.4826 * MAD_{q!=p}(L_iq) )

    Step 8 is the across-batch step: q ranges over the OTHER proteins carrying the
    same compound i, and because each protein sits in exactly one batch, that
    reference set spans batches (11,871 of 12,399 compounds appear in all 17).
    Both the median and the MAD are leave-one-out -- the row being scored is
    excluded from its own reference, as written on the slide.

    THE MAD FLOOR (MAD_MIN). Step 8 as written is unusable on censored data,
    because MAD_{q!=p} collapses toward zero whenever a majority of the other
    proteins share one identical L -- which is the norm here, since whole batches
    float together (all 8 proteins at the 3000 floor => every E = 1 => every L = 0).
    Two failures followed from that, and they are the same failure:

      * MAD exactly 0 -> 0/0 -> no score. On 20to22 that was 43.8% of rows, and it
        hit 100% of the compounds detected on a single protein -- i.e. it threw away
        precisely the selective binders the screen is looking for (XS837730b on FDPS,
        168k/220k/156k in triplicate, fold 60, BINARY_LABEL=1, scored NaN).
      * MAD barely above 0 -> a colossal score. A BRD1 row whose three replicates
        were ALL 3000 came top of its file at 58.7, because its reference MAD was
        0.0019 -- that number is the gap between two batch scale factors, not
        biological spread.

    So the denominator is floored: 1.4826 * max(MAD, MAD_MIN), where MAD_MIN is
    ESTIMATED FROM THIS DATASET by estimate_mad_min() -- not hard-coded, because it
    depends on how many proteins each compound is carried by. Pass `mad_min=` to
    override. On 20to22 the estimate is ~0.12, which sends the BRD1 artifact to 0.93
    and gives XS837730b 32.7.

    Note the trade this makes: for a clamped row the score is no longer a true
    robust z, it is deviation / fixed scale (the standard variance-moderation
    trade, as in limma). `ns1_ref_mad` carries the RAW unclamped MAD so any row can
    be audited or re-filtered without recomputing.

    Returns a dict of columns plus the mad_min actually used. S is NaN only for a
    compound carried by fewer than 3 proteins.
    """
    n = len(batch)
    ubatch = sorted(set(batch))

    # 1-4. batch medians on DETECTED values only, then the median of those medians
    bmed = {}
    for b in ubatch:
        sub = reps[batch == b]
        det = sub[sub > FLOOR]
        bmed[b] = float(np.median(det)) if det.size else np.nan
    ref = float(np.median([bmed[b] for b in ubatch]))
    scale = {b: ref / bmed[b] for b in ubatch}
    print(f"  reference (median of {len(ubatch)} batch medians) = {ref:,.1f}")
    for b in ubatch:
        print(f"    {b:<14} median {bmed[b]:>12,.1f}   scale {scale[b]:>7.4f}")

    # 5-6. scale everything (floor included), then collapse the replicates
    srow = np.array([scale[b] for b in batch], dtype=float)
    M = np.median(reps * srow[:, None], axis=1)

    # 7. within-batch enrichment: target vs the MEAN of the same-batch non-targets.
    #    A mean, not a median, deliberately: a median over the 7 background proteins
    #    would discard a floored minority entirely (3 floored + 4 strong -> the
    #    median is a strong value and the floors carry no weight), whereas the mean
    #    lets every background protein count. The median in this chain acts only
    #    across the 3 replicates of a single row.
    grp = pd.Series(M).groupby([pd.Series(batch), pd.Series(comp)])
    gsum = grp.transform("sum").to_numpy()
    gcnt = grp.transform("size").to_numpy()
    with np.errstate(divide="ignore", invalid="ignore"):
        nt = (gsum - M) / (gcnt - 1)
        nt[gcnt < 2] = np.nan
        E = M / nt
        L = np.log2(E)

    # 7b. ACROSS-batch enrichment, on the very same scaled values. Identical to
    #     step 7 except the background is every other protein carrying this
    #     compound, all batches pooled -- which is only defensible because step 5
    #     already put the batches on a common scale. Cross-batch pooling of RAW
    #     intensities is what the batch effect ruins; after scaling it is legitimate.
    cgrp = pd.Series(M).groupby(pd.Series(comp))
    csum = cgrp.transform("sum").to_numpy()
    ccnt = cgrp.transform("size").to_numpy()
    with np.errstate(divide="ignore", invalid="ignore"):
        nt_across = (csum - M) / (ccnt - 1)
        nt_across[ccnt < 2] = np.nan
        E_across = M / nt_across

    # 8a. leave-one-out reference median and MAD over the other proteins of the
    #     same compound. Done in its own pass because the MAD floor is estimated
    #     FROM these values, so it cannot be applied while they are being built.
    ref_med = np.full(n, np.nan)
    ref_mad = np.full(n, np.nan)
    n_small = 0
    for _, idx in pd.Series(comp).groupby(pd.Series(comp)).indices.items():
        Lg = L[idx]
        k = len(Lg)
        if k < 3:
            n_small += k
            continue
        wo = np.broadcast_to(Lg, (k, k))[~np.eye(k, dtype=bool)].reshape(k, k - 1)
        if np.isnan(wo).any():
            ref_med[idx] = np.nanmedian(wo, axis=1)
            ref_mad[idx] = np.nanmedian(np.abs(wo - ref_med[idx][:, None]), axis=1)
        else:
            ref_med[idx] = np.median(wo, axis=1)
            ref_mad[idx] = np.median(np.abs(wo - ref_med[idx][:, None]), axis=1)

    # 8b. the MAD floor -- re-derived for THIS dataset unless overridden
    detected = (reps > FLOOR).any(axis=1)
    if mad_min is None:
        mad_min, _ = estimate_mad_min(ref_mad, detected, comp)
    else:
        mad_min = float(mad_min)
        print(f"  MAD_MIN overridden by caller = {mad_min:.4f}")

    # 8c. the robust Z itself
    with np.errstate(divide="ignore", invalid="ignore"):
        S = (L - ref_med) / (1.4826 * np.maximum(ref_mad, mad_min))
    S[~np.isfinite(S)] = np.nan
    n_clamped = int(np.sum(ref_mad < mad_min))
    print(f"  Normalized score1: {np.isfinite(S).sum():,} finite / {n:,} rows "
          f"| MAD floored on {n_clamped:,} rows "
          f"({n_clamped / max(n - n_small, 1):.1%}) "
          f"| {n_small:,} rows in <3-protein groups")
    print(f"  ns1_enrichment_across: median {np.nanmedian(E_across):.3f} "
          f"| p99 {np.nanpercentile(E_across, 99):,.1f} "
          f"| {np.isnan(E_across).sum():,} NaN")
    return {"scale": srow, "M": M, "nt_within": nt, "E_within": E, "L": L, "S": S,
            "ref_med": ref_med, "ref_mad": ref_mad,
            "nt_across": nt_across, "E_across": E_across, "mad_min": mad_min}


def bead_columns(comp, tmed, beads_path):
    """bead_signal / bead_detected / bead_ratio / bead_binder_flag_r5|r10.

    Returns all-NaN/zero columns (with a warning) if the workbook is absent, so a
    run without the bead reference still produces every other column.
    """
    n = len(comp)
    if not beads_path or not os.path.exists(beads_path):
        print(f"  [warn] bead workbook not found ({beads_path}) -> bead columns left empty")
        nan = np.full(n, np.nan)
        z = np.zeros(n, dtype=int)
        return nan, z, nan, z, z
    b = pd.read_excel(beads_path)
    b["bead_signal"] = np.maximum(b[C5].median(axis=1), b[C10].median(axis=1))
    bead = (b[["SGC ID for Component", "bead_signal"]]
            .rename(columns={"SGC ID for Component": "COMPOUND_ID"})
            .dropna(subset=["COMPOUND_ID"]).drop_duplicates("COMPOUND_ID"))
    bmap = dict(zip(bead["COMPOUND_ID"], bead["bead_signal"]))
    bead_signal = np.array([bmap.get(c, np.nan) for c in comp])
    bead_detected = (bead_signal > FLOOR).astype(int)
    bead_ratio = tmed / bead_signal
    bbr5 = ((bead_signal > FLOOR) & (bead_ratio < 5)).astype(int)
    bbr10 = ((bead_signal > FLOOR) & (bead_ratio < 10)).astype(int)
    return bead_signal, bead_detected, bead_ratio, bbr5, bbr10


def compute_columns(paths, beads_path, do_validate=False, mad_min=None):
    """Every appended column, for all files at once. Returns a DataFrame + `_src`."""
    print(f"Loading {len(paths)} files (lightweight) ...")
    parts = []
    for p in paths:
        want = ["COMPOUND_ID", "ASMS_BATCH_NAME"] + REP
        head = pd.read_csv(p, nrows=0).columns
        if "n_detected" in head:
            want.append("n_detected")
        d = pd.read_csv(p, usecols=want, low_memory=False)
        d["_src"] = os.path.basename(p)
        parts.append(d)
    big = pd.concat(parts, ignore_index=True)

    reps_raw = big[REP].to_numpy(dtype=float)
    n = len(big)
    batch = big["ASMS_BATCH_NAME"].to_numpy()
    comp = big["COMPOUND_ID"].to_numpy()
    print(f"{n:,} rows | {big['COMPOUND_ID'].nunique():,} compounds | "
          f"{len(set(batch))} batches")

    # ---------- 0. article equation 1 (raw replicates, missing EXCLUDED) -------
    print("article within-batch enrichment fold (equation 1) ...")
    fold_article = article_fold_within(big, reps_raw)

    # ---------- floor imputation for everything below -------------------------
    n_nan_rows = int(np.isnan(reps_raw).any(axis=1).sum())
    reps = np.where(np.isnan(reps_raw), FLOOR, reps_raw)
    if n_nan_rows:
        print(f"imputed {n_nan_rows} rows with a NaN replicate -> floor {FLOOR:g} "
              f"(median/norm/pctnorm/bead columns only)")

    ndet = (big["n_detected"].to_numpy() if "n_detected" in big.columns
            else (reps > FLOOR).sum(axis=1))
    tmed = np.median(reps, axis=1)

    idx_by_comp = defaultdict(list)
    for i, c in enumerate(comp):
        idx_by_comp[c].append(i)

    # ---------- 1. raw median enrichment (within + across) --------------------
    print("raw median enrichment (within + across Mann-Whitney) ...")
    nt_w, nt_a, p_w, p_a = enrichment_and_p(reps, tmed, idx_by_comp, batch, ndet)
    enr_w, enr_a = tmed / nt_w, tmed / nt_a

    # ---------- 2. median normalization ---------------------------------------
    print("median-normalized (across only; within derived) ...")
    grand = np.median(reps[reps > FLOOR])
    frow = np.empty(n)
    for b in set(batch):
        m = batch == b
        frow[m] = grand / np.median(reps[m][reps[m] > FLOOR])
    reps_n = reps * frow[:, None]
    tmed_n = tmed * frow
    _, nt_a_n, _, p_a_n = enrichment_and_p(reps_n, tmed_n, idx_by_comp, batch, ndet,
                                           do_within=False)
    enr_a_n = tmed_n / nt_a_n
    nt_w_n = nt_w * frow                 # within background scales linearly
    enr_w_n = enr_w                      # ratio unchanged
    p_w_n = p_w                          # rank unchanged

    # ---------- 3. percentile normalization -----------------------------------
    print("percentile-normalized (rank -> pooled reference) ...")
    pooled_sorted = np.sort(reps.ravel())
    grid = np.linspace(0.0, 1.0, len(pooled_sorted))
    reps_p = np.empty_like(reps)
    for b in set(batch):
        m = batch == b
        sub = reps[m].ravel()
        pct = rankdata(sub, method="average") / len(sub)
        reps_p[m] = np.interp(pct, grid, pooled_sorted).reshape(-1, len(REP))
    tmed_p = np.median(reps_p, axis=1)
    # within p-value is recomputed here, NOT copied from the raw pass: the
    # transform's plateau at the floor creates ties that change the MW ranks
    nt_w_p, nt_a_p, p_w_p, p_a_p = enrichment_and_p(reps_p, tmed_p, idx_by_comp,
                                                    batch, ndet)
    enr_w_p, enr_a_p = tmed_p / nt_w_p, tmed_p / nt_a_p

    # the missing cell of the 2x2: quantile scaling with a MEAN across-batch
    # background (enrichment_and_p only ever builds a median one). Same pooled
    # reference set -- every other protein carrying this compound, all batches.
    _cg = pd.Series(tmed_p).groupby(pd.Series(comp))
    _csum = _cg.transform("sum").to_numpy()
    _ccnt = _cg.transform("size").to_numpy()
    with np.errstate(divide="ignore", invalid="ignore"):
        nt_a_p_mean = (_csum - tmed_p) / (_ccnt - 1)
        nt_a_p_mean[_ccnt < 2] = np.nan
        enr_a_p_mean = tmed_p / nt_a_p_mean

    # ---------- 4. bead specificity -------------------------------------------
    print("bead specificity ...")
    bead_signal, bead_detected, bead_ratio, bbr5, bbr10 = bead_columns(
        comp, tmed, beads_path)

    # ---------- 5. Normalized score1 ------------------------------------------
    print("Normalized score1 (batch scaling -> within-batch E -> robust Z) ...")
    ns1 = normalized_score1(reps, batch, comp, mad_min=mad_min)

    newdf = pd.DataFrame({
        "fold_articlebased_within": fold_article,
        "target_median": tmed, "signal_flag": signal_flags(reps),
        "nontarget_median_within": nt_w, "enrichment_within": enr_w,
        "pvalue_within": p_w, "label_within": labels(enr_w, p_w),
        "nontarget_median_across": nt_a, "enrichment_across": enr_a,
        "pvalue_across": p_a, "label_across": labels(enr_a, p_a),
        "norm_scale_factor": frow, "target_median_norm": tmed_n,
        "nontarget_median_within_norm": nt_w_n, "enrichment_within_norm": enr_w_n,
        "pvalue_within_norm": p_w_n, "label_within_norm": labels(enr_w_n, p_w_n),
        "nontarget_median_across_norm": nt_a_n, "enrichment_across_norm": enr_a_n,
        "pvalue_across_norm": p_a_n, "label_across_norm": labels(enr_a_n, p_a_n),
        "target_median_pctnorm": tmed_p,
        "nontarget_median_within_pctnorm": nt_w_p,
        "enrichment_within_pctnorm": enr_w_p,
        "pvalue_within_pctnorm": p_w_p, "label_within_pctnorm": labels(enr_w_p, p_w_p),
        "nontarget_median_across_pctnorm": nt_a_p,
        "enrichment_across_pctnorm": enr_a_p,
        "pvalue_across_pctnorm": p_a_p, "label_across_pctnorm": labels(enr_a_p, p_a_p),
        "bead_signal": bead_signal, "bead_detected": bead_detected,
        "bead_ratio": bead_ratio, "bead_binder_flag_r5": bbr5,
        "bead_binder_flag_r10": bbr10,
        "ns1_scale_factor": ns1["scale"], "ns1_target_median": ns1["M"],
        "ns1_nontarget_mean_within": ns1["nt_within"],
        "ns1_enrichment_within": ns1["E_within"],
        "ns1_nontarget_mean_across": ns1["nt_across"],
        "ns1_enrichment_across": ns1["E_across"],
        "ns1_log2_enrichment": ns1["L"],
        "ns1_ref_median": ns1["ref_med"], "ns1_ref_mad": ns1["ref_mad"],
        # the across-batch 2x2 under explicit names (see NEW_COLS)
        "enrich_across_medscale_medbg": enr_a_n,
        "enrich_across_medscale_meanbg": ns1["E_across"],
        "enrich_across_quantscale_medbg": enr_a_p,
        "enrich_across_quantscale_meanbg": enr_a_p_mean,
        "Normalized score1": ns1["S"],
    })[NEW_COLS]

    if do_validate:
        validate(reps, reps_raw, tmed, comp, batch, ndet, nt_a, p_a, fold_article)

    for c in ["nontarget_median_across", "nontarget_median_across_norm",
              "nontarget_median_across_pctnorm", "fold_articlebased_within"]:
        print(f"  NaN in {c}: {newdf[c].isna().sum():,} ({newdf[c].isna().mean():.1%})")
    for lab in ["label_within", "label_across", "label_across_norm",
                "label_across_pctnorm"]:
        print(f"  {lab}=1: {int(newdf[lab].sum()):,}")
    hits = (newdf["label_within"] == 1)
    print(f"  bead_binder_flag_r10=1 among label_within hits: "
          f"{int((hits & (newdf['bead_binder_flag_r10'] == 1)).sum()):,}")

    # which of the across-batch 2x2 actually flattens the batch effect? A metric
    # that has removed it should have the SAME median in every batch, so max/min
    # over the per-batch medians is the score to minimise (1.00 = perfect).
    print("\n  across-batch enrichment, batch flatness (max/min of per-batch "
          "median; 1.00 = batch-free):")
    _cmp = ["enrichment_across"] + [c for c in NEW_COLS if c.startswith("enrich_across_")]
    _rows = []
    for c in _cmp:
        med = pd.Series(newdf[c].to_numpy()).groupby(pd.Series(batch)).median()
        _rows.append((c, float(med.max() / med.min()) if med.min() > 0 else np.nan))
    for c, r in sorted(_rows, key=lambda t: (np.isnan(t[1]), t[1])):
        tag = "  <- unscaled baseline" if c == "enrichment_across" else ""
        print(f"    {r:>6.3f}   {c}{tag}")

    newdf["_src"] = big["_src"].to_numpy()
    return newdf


def validate(reps, reps_raw, tmed, comp, batch, ndet, nt_a, p_a, fold_article,
             n_check=200, seed=0):
    """Direct recompute of the across background, its MWU, and equation 1."""
    rng = np.random.default_rng(seed)
    idx_by_comp = defaultdict(list)
    for i, c in enumerate(comp):
        idx_by_comp[c].append(i)
    idx = rng.choice(len(reps), size=min(n_check, len(reps)), replace=False)
    A = peak_intensity(reps_raw)
    max_nt = max_p = max_fold = 0.0
    for i in idx:
        peers = np.array([k for k in idx_by_comp[comp[i]] if k != i])
        if not peers.size:
            continue
        max_nt = max(max_nt, abs(np.median(tmed[peers]) - nt_a[i]))
        if ndet[i] > 0:
            pv = mannwhitneyu(reps[i], tmed[peers], alternative="two-sided").pvalue
            max_p = max(max_p, abs(pv - p_a[i]))
        # equation 1: same batch, measured peers only
        bg = np.array([A[k] for k in peers
                       if batch[k] == batch[i] and not np.isnan(A[k])])
        if bg.size and not np.isnan(A[i]):
            expect = A[i] * bg.size / bg.sum()
            max_fold = max(max_fold, abs(expect - fold_article[i]) / max(expect, 1e-12))
    print(f"[validate] {len(idx)} rows | max abs-diff nt_across={max_nt:.2e} "
          f"p_across={max_p:.2e} | max rel-diff fold_articlebased_within={max_fold:.2e}")


def write_outputs(paths, newdf, out_dir):
    """Re-read each source file, append the computed columns, write to out_dir.

    Resilient to locked output files (open in Excel): skips them and returns the
    names that could not be written.
    """
    src = newdf["_src"].to_numpy()
    locked = []
    for p in paths:
        name = os.path.basename(p)
        add = newdf.loc[src == name, NEW_COLS].reset_index(drop=True)
        try:
            full = pd.read_csv(p, low_memory=False)
            for c in NEW_COLS:
                full[c] = add[c].to_numpy()
            full.to_csv(os.path.join(out_dir, name), index=False)
        except PermissionError:
            locked.append(name)
    print(f"\nwrote {len(paths) - len(locked)}/{len(paths)} files to {out_dir}/")
    if locked:
        print(f"[LOCKED] {len(locked)} not written (open in Excel?): {locked}")
        print("  close them, then re-run with --write-only")
    return locked


# --------------------------------------------------------------------------- #
#  molecule x protein matrix (shared by the heatmap and the CSV/XLSX export)
# --------------------------------------------------------------------------- #

class Matrix:
    """Hit-molecule x protein median-intensity matrix, rows sorted by floor count.

    Rows    : molecules with BINARY_LABEL=1 for at least one protein.
    Columns : every protein, ordered by hit count descending.
    Cells   : median of the 3 replicates; not-measured -> the 3000 floor, so the
              heatmap and the exported matrix agree cell for cell.
    Sorting : by n_floor_3000_of_median descending (most floors on top).
    """

    def __init__(self, out_dir):
        self.files = sorted(glob.glob(os.path.join(out_dir, "*.csv")))
        self.proteins = [os.path.basename(f)[:-4] for f in self.files]
        self._pass1()
        self._pass2()
        self._build()

    def _pass1(self):
        """hit compounds, SMILES, per-protein hit count, PROTEIN_NAME."""
        self.hit_compounds = set()
        self.smiles = {}
        self.prot_hits = {}
        self.protein_name = {}
        for f in self.files:
            prot = os.path.basename(f)[:-4]
            head = pd.read_csv(f, nrows=0).columns
            want = [c for c in ["COMPOUND_ID", "SMILES", "BINARY_LABEL",
                                "PROTEIN_NAME"] if c in head]
            d = pd.read_csv(f, usecols=want, low_memory=False)
            if "SMILES" in d:
                for c, s in zip(d["COMPOUND_ID"], d["SMILES"]):
                    self.smiles.setdefault(str(c).strip(), "" if pd.isna(s) else str(s))
            hit = d["BINARY_LABEL"].astype(str).str.strip() == "1"
            self.hit_compounds.update(d.loc[hit, "COMPOUND_ID"].astype(str).str.strip())
            self.prot_hits[prot] = int(hit.sum())
            self.protein_name[prot] = (str(d["PROTEIN_NAME"].iloc[0])
                                       if "PROTEIN_NAME" in d and len(d) else prot)
        print(f"hit compounds: {len(self.hit_compounds)} | proteins: {len(self.files)}")
        self.prot_order = sorted(self.proteins, key=lambda p: -self.prot_hits[p])
        self.prank = {p: i for i, p in enumerate(self.prot_order)}

    def _pass2(self):
        """medians, replicates, hit flags, SPR flags and the appended columns."""
        self.median_v = defaultdict(dict)
        self.reps_v = defaultdict(dict)
        self.is_hit = {}
        self.has_spr = {}
        self.hits_of = defaultdict(list)
        for f in self.files:
            prot = os.path.basename(f)[:-4]
            head = pd.read_csv(f, nrows=0).columns
            self.extra_cols = [c for c in NEW_COLS if c in head]
            want = ([c for c in ["COMPOUND_ID", "BINARY_LABEL", "SPR_CATEGORY"]
                     if c in head] + REP + self.extra_cols)
            d = pd.read_csv(f, usecols=want, low_memory=False)
            d["COMPOUND_ID"] = d["COMPOUND_ID"].astype(str).str.strip()
            d = d[d["COMPOUND_ID"].isin(self.hit_compounds)]
            reps = d[REP].to_numpy(dtype=float)
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", RuntimeWarning)
                med = np.nanmedian(reps, axis=1)
            hit = d["BINARY_LABEL"].astype(str).str.strip() == "1"
            spr = (d["SPR_CATEGORY"].astype(str).str.strip()
                   .isin(["not_found", "", "nan"]).map(lambda x: not x)
                   if "SPR_CATEGORY" in d else pd.Series(False, index=d.index))
            for k, (comp, m, h, s) in enumerate(zip(d["COMPOUND_ID"], med, hit, spr)):
                self.reps_v[comp][prot] = reps[k].tolist()
                if np.isfinite(m):
                    self.median_v[comp][prot] = m
                self.is_hit[(comp, prot)] = bool(h)
                self.has_spr[(comp, prot)] = bool(s)
            for comp, row in zip(d.loc[hit, "COMPOUND_ID"],
                                 d.loc[hit, self.extra_cols].astype(str).to_dict("records")):
                self.hits_of[comp].append((self.prank[prot], prot, row))

    def _build(self):
        compounds = sorted(self.hit_compounds)
        po = self.prot_order
        Mmed = np.array([[self.median_v.get(c, {}).get(p, np.nan) for p in po]
                         for c in compounds], dtype=float)
        self.n_missing = int((~np.isfinite(Mmed)).sum())
        Mmed[~np.isfinite(Mmed)] = FLOOR

        def triplet(c, p):
            t = self.reps_v.get(c, {}).get(p)
            if t is None:
                return [FLOOR] * len(REP)
            return [FLOOR if (x is None or not np.isfinite(x)) else x for x in t]

        Mrep = np.array([[v for p in po for v in triplet(c, p)] for c in compounds],
                        dtype=float)
        n_floor_med = (Mmed <= FLOOR).sum(axis=1)
        n_floor_rep = (Mrep <= FLOOR).sum(axis=1)

        order = np.argsort(-n_floor_med, kind="stable")   # most floors on top
        self.compounds = [compounds[i] for i in order]
        self.Mmed = Mmed[order]
        self.Mrep = Mrep[order]
        self.n_floor_med = n_floor_med[order]
        self.n_floor_rep = n_floor_rep[order]
        self.cidx = {c: i for i, c in enumerate(self.compounds)}
        print(f"matrix {len(self.compounds)} molecules x {len(po)} proteins; "
              f"missing->floor={self.n_missing}; floor-count "
              f"top={self.n_floor_med[0]} .. bottom={self.n_floor_med[-1]} of {len(po)}")


def plot_heatmap(mx, png_path, title_tag):
    """Median-intensity heatmap: red outline = ASMS hit, white dot = SPR-tested."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.colors import LogNorm
    from matplotlib.patches import Rectangle

    M = mx.Mmed
    nR, nC = M.shape
    vmin, vmax_full = np.nanmin(M), np.nanmax(M)
    vmax = np.nanpercentile(M, 99.9)
    fig, ax = plt.subplots(figsize=(16, max(8, nR * 0.026)))
    im = ax.imshow(M, aspect="auto", cmap=plt.get_cmap("viridis").copy(),
                   norm=LogNorm(vmin=max(vmin, 1), vmax=vmax), interpolation="nearest")
    for (comp, prot), h in mx.is_hit.items():
        if h and comp in mx.cidx:
            ax.add_patch(Rectangle((mx.prank[prot] - 0.5, mx.cidx[comp] - 0.5), 1, 1,
                                   fill=False, edgecolor="red", linewidth=0.7))
    sx = [mx.prank[p] for (c, p), s in mx.has_spr.items() if s and c in mx.cidx]
    sy = [mx.cidx[c] for (c, p), s in mx.has_spr.items() if s and c in mx.cidx]
    ax.scatter(sx, sy, s=3, c="white", marker=".", linewidths=0)
    ax.set_xticks(range(nC)); ax.set_xticklabels(mx.prot_order, rotation=90, fontsize=6)
    ax.set_yticks([])
    ax.set_ylabel(f"{nR} molecules - sorted by #{FLOOR:.0f}(floor) cells (most floors on top)")
    ax.set_xlabel(f"Protein target (all {nC})")
    ax.set_title(f"{title_tag}: ASMS median intensity matrix - rows sorted by number of "
                 f"{FLOOR:.0f}(floor) values\ntop = most floors (least signal) -> bottom = "
                 f"fewest (most signal) | viridis dark=floor -> yellow=high | "
                 f"red outline = ASMS hit | white dot = SPR-tested")
    cb = fig.colorbar(im, ax=ax, fraction=0.02, pad=0.01)
    cb.set_label("ASMS median intensity (log)")
    fig.tight_layout()
    fig.savefig(png_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"saved {png_path}  ({vmin:.0f}..{vmax_full:.0f}, color cap {vmax:.0f})")


def export_matrix(mx, csv_path, xlsx_path):
    """The same matrix as a wide CSV + a colour-coded XLSX.

    Left block: identity, floor counts, then every appended analysis column taken
    from the file(s) where the molecule is a hit -- comma-joined in protein-rank
    order, so position i of each list refers to the same hit protein.
    """
    from openpyxl.styles import PatternFill

    pname_col, extra = [], {c: [] for c in mx.extra_cols}
    for c in mx.compounds:
        hits = sorted(mx.hits_of.get(c, []), key=lambda t: t[0])
        pname_col.append(", ".join(mx.protein_name.get(p, p) for _, p, _ in hits))
        for ec in mx.extra_cols:
            extra[ec].append(", ".join(v.get(ec, "") for _, _, v in hits))

    meta = {
        "COMPOUND_ID": mx.compounds,
        "SMILES": [mx.smiles.get(c, "") for c in mx.compounds],
        "PROTEIN_NAME": pname_col,
        "n_floor_3000_of_median": mx.n_floor_med,
        "n_floor_3000_in_replicas": mx.n_floor_rep,
    }
    meta.update(extra)
    median_headers = [f"{p} (median)" for p in mx.prot_order]
    rep_headers = [f"{p}_{rc.replace('POS_INT_', '')}" for p in mx.prot_order for rc in REP]
    df = pd.concat([pd.DataFrame(meta).reset_index(drop=True),
                    pd.DataFrame(mx.Mmed, columns=median_headers),
                    pd.DataFrame(mx.Mrep, columns=rep_headers)], axis=1)

    base = ["COMPOUND_ID", "SMILES", "PROTEIN_NAME",
            "n_floor_3000_of_median", "n_floor_3000_in_replicas"]
    enr_p = [c for c in mx.extra_cols
             if c.startswith(("enrichment", "pvalue", "fold_"))]
    bead = [c for c in mx.extra_cols if c.startswith("bead")]
    placed = set(enr_p) | {"signal_flag"} | set(bead)
    other = [c for c in mx.extra_cols if c not in placed]
    df = df[base + enr_p + [c for c in ["signal_flag"] if c in mx.extra_cols]
            + bead + other + median_headers + rep_headers]

    df.to_csv(csv_path, index=False)
    print(f"saved {csv_path}")

    fills = {}
    for c in bead:
        fills[c] = PatternFill("solid", fgColor="D9D9D9")     # gray
    for c in median_headers:
        fills[c] = PatternFill("solid", fgColor="DDEBF7")     # light blue
    for c in rep_headers:
        fills[c] = PatternFill("solid", fgColor="E2EFDA")     # light green
    path = xlsx_path
    for attempt in range(5):
        try:
            with pd.ExcelWriter(path, engine="openpyxl") as xw:
                df.to_excel(xw, sheet_name="matrix_sorted_floor", index=False)
                ws = xw.sheets["matrix_sorted_floor"]
                for j, cname in enumerate(df.columns, start=1):
                    fill = fills.get(cname)
                    if fill is None:
                        continue
                    for i in range(1, df.shape[0] + 2):       # header + data rows
                        ws.cell(row=i, column=j).fill = fill
            break
        except PermissionError:
            path = xlsx_path.replace(".xlsx", f"_{attempt + 2}.xlsx")
            print(f"  (locked - retrying as {os.path.basename(path)})")
    print(f"saved {path}")
    n_meta = df.shape[1] - len(median_headers) - len(rep_headers)
    print(f"  shape: {df.shape[0]} molecules x {df.shape[1]} columns "
          f"({len(median_headers)} median + {len(rep_headers)} replica + {n_meta} meta)")
    multi = sum(1 for c in mx.compounds if len(mx.hits_of.get(c, [])) > 1)
    print(f"  molecules hit in >1 file (comma-joined analysis columns): {multi}")


# --------------------------------------------------------------------------- #

def parse_args(argv=None):
    ap = argparse.ArgumentParser(
        description="Full multi-batch analysis: all computed columns + heatmap + matrix.")
    ap.add_argument("--input-dir", required=True,
                    help="folder of raw per-protein CSVs, e.g. MultiBatch_20to36")
    ap.add_argument("--out-dir", default=None,
                    help="default: <input-dir>_modified")
    ap.add_argument("--results-dir", default=None,
                    help="default: <input-dir>_results (plots, matrix, cache)")
    ap.add_argument("--beads", default=os.path.join("MultiBatchResults", "beads_clean.xlsx"),
                    help="bead reference workbook")
    ap.add_argument("--write-only", action="store_true",
                    help="re-write the _modified folder from the cached columns")
    ap.add_argument("--skip-columns", action="store_true",
                    help="keep the existing _modified folder, only rebuild plots/matrix")
    ap.add_argument("--skip-heatmap", action="store_true")
    ap.add_argument("--skip-matrix", action="store_true")
    ap.add_argument("--validate", action="store_true")
    ap.add_argument("--mad-min", type=float, default=None,
                    help="MAD floor for 'Normalized score1', in log2 units. "
                         "Omit to RE-DERIVE it from the dataset (the default: the "
                         "5th percentile of the leave-one-out MAD among compounds "
                         "detected on >=75%% of their proteins). Pass a number to "
                         "pin it, e.g. --mad-min 0.12 to reproduce an earlier run.")
    return ap.parse_args(argv)


def main(argv=None):
    a = parse_args(argv)
    src = a.input_dir.rstrip("/\\")
    name = os.path.basename(src)
    out_dir = a.out_dir or f"{src}_modified"
    res_dir = a.results_dir or f"{src}_results"
    cache = os.path.join(res_dir, "_newcols.pkl")

    if not os.path.isdir(src):
        sys.exit(f"--input-dir not found: {src}")
    paths = sorted(glob.glob(os.path.join(src, "*.csv")))
    if not paths:
        sys.exit(f"no CSV files in {src}/")
    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(res_dir, exist_ok=True)
    print(f"input   : {src}/ ({len(paths)} files)\noutput  : {out_dir}/\n"
          f"results : {res_dir}/\n")

    if not a.skip_columns:
        if a.write_only:
            print(f"write-only: loading cached columns from {cache}")
            newdf = pd.read_pickle(cache)
        else:
            newdf = compute_columns(paths, a.beads, a.validate, a.mad_min)
            # cache BEFORE writing, so one locked output file never forces a
            # recompute -- resume with --write-only after closing it
            newdf.to_pickle(cache)
            print(f"cached computed columns to {cache}")
        write_outputs(paths, newdf, out_dir)

    if a.skip_heatmap and a.skip_matrix:
        return
    print("\nbuilding molecule x protein matrix ...")
    mx = Matrix(out_dir)
    if not a.skip_heatmap:
        plot_heatmap(mx, os.path.join(res_dir, f"{name}_hit_matrix_intensity_sorted_floor.png"),
                     name)
    if not a.skip_matrix:
        export_matrix(mx,
                      os.path.join(res_dir, f"{name}_matrix_sorted_floor_full.csv"),
                      os.path.join(res_dir, f"{name}_matrix_sorted_floor_full.xlsx"))
    print("\ndone.")


if __name__ == "__main__":
    main()
