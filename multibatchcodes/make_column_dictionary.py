# -*- coding: utf-8 -*-
"""Write the column dictionary for run_multibatch_pipeline.py to an Excel file.

Three sheets:
    Added columns     the 35 columns the pipeline appends to every row
    Matrix export     the extra columns of <name>_matrix_sorted_floor_full.csv/xlsx
    Parameters        thresholds, notation and the missing-value conventions

Run from the repo root:
    python multibatchcodes/make_column_dictionary.py
    python multibatchcodes/make_column_dictionary.py --out some/other/file.xlsx
"""

import os
import argparse

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DEFAULT_OUT = os.path.join("MultiBatch_20to36_plots",
                           "multibatch_pipeline_column_definitions.xlsx")

# A = mean of the replicates, M = median of the replicates, f = norm_scale_factor
ADDED = [
    # (column, group, definition, formula, background / scope, notes)
    ("fold_articlebased_within", "Article equation 1",
     "Enrichment fold of the compound on this protein, using the other proteins of the "
     "same batch as the control. This is equation (1) of the E-ASMS paper.",
     "fold = A_POI x N / sum(A_i)   [equivalent: A_POI / mean(A_i)]",
     "Same compound, the other proteins of the SAME ASMS_BATCH_NAME (N = 7 when the "
     "batch is complete).",
     "A = mean(REP1..3). N counts only proteins with a measured intensity. NaN if this "
     "protein has no measurement. Reproduces the pipeline's own fold_change."),

    ("target_median", "Per molecule",
     "Median ASMS peak intensity of this compound on this protein.",
     "M = median(REP1, REP2, REP3)",
     "This row only.",
     "Missing replicates are set to the 3000 floor first."),
    ("signal_flag", "Per molecule",
     "Verbal strength of the raw signal: how many replicates are strong / moderate.",
     "'<k> strong' if k reps >= 1e6; else '<k> moderate' if k reps in [1e4, 1e6); "
     "else 'none'",
     "This row only.",
     "Strong takes precedence over moderate."),

    ("nontarget_median_within", "Median enrichment - within batch",
     "Background level of this compound inside its own batch: the median of the other "
     "proteins' median intensities.",
     "median over the other same-batch proteins of their target_median",
     "Same compound, other proteins of the same batch (7 when complete).",
     "Per-protein median first, then median across proteins."),
    ("enrichment_within", "Median enrichment - within batch",
     "How much more of the compound this protein pulled down than its batch background.",
     "enrichment_within = target_median / nontarget_median_within",
     "As above.", ""),
    ("pvalue_within", "Median enrichment - within batch",
     "Significance of that difference.",
     "Mann-Whitney U, two-sided: this row's 3 replicates vs the same-batch background "
     "medians",
     "As above.",
     "Forced to 1.0 when n_detected = 0 (never detected)."),
    ("label_within", "Median enrichment - within batch",
     "Hit call from the within-batch statistics.",
     "1 if enrichment_within >= 5 AND pvalue_within <= 0.05, else 0",
     "As above.", "Paper's thresholds."),

    ("nontarget_median_across", "Median enrichment - across batches",
     "Background level of this compound over every other protein in the folder.",
     "median over ALL other proteins of their target_median",
     "Same compound, all other proteins, all batches pooled.",
     "Wider, less batch-sensitive background than the within-batch one."),
    ("enrichment_across", "Median enrichment - across batches",
     "Enrichment against that wider background.",
     "enrichment_across = target_median / nontarget_median_across", "As above.", ""),
    ("pvalue_across", "Median enrichment - across batches",
     "Significance against the wider background.",
     "Mann-Whitney U, two-sided: 3 replicates vs all other proteins' medians",
     "As above.", "Forced to 1.0 when n_detected = 0."),
    ("label_across", "Median enrichment - across batches",
     "Hit call from the across-batch statistics.",
     "1 if enrichment_across >= 5 AND pvalue_across <= 0.05, else 0", "As above.", ""),

    ("norm_scale_factor", "Median-normalized (_norm)",
     "Per-batch scaling that puts every batch on a common intensity level.",
     "f = median(all above-floor values, all batches) / median(above-floor values of "
     "THIS batch)",
     "One value per batch; identical for every row of a batch.",
     "Corrects batch-to-batch instrument response."),
    ("target_median_norm", "Median-normalized (_norm)",
     "This row's median intensity after batch scaling.",
     "target_median x f", "This row.", ""),
    ("nontarget_median_within_norm", "Median-normalized (_norm)",
     "Within-batch background after scaling.",
     "nontarget_median_within x f", "Same-batch proteins.",
     "Same factor for the whole batch, so it is a pure rescale."),
    ("enrichment_within_norm", "Median-normalized (_norm)",
     "Within-batch enrichment after scaling.",
     "= enrichment_within (unchanged)", "Same-batch proteins.",
     "The scaling is linear and per batch, so the ratio cancels exactly."),
    ("pvalue_within_norm", "Median-normalized (_norm)",
     "Within-batch p-value after scaling.",
     "= pvalue_within (unchanged)", "Same-batch proteins.",
     "A strictly increasing rescale cannot change ranks, so the rank test is identical."),
    ("label_within_norm", "Median-normalized (_norm)",
     "Hit call on the normalized within-batch statistics.",
     "1 if enrichment_within_norm >= 5 AND pvalue_within_norm <= 0.05, else 0",
     "Same-batch proteins.", ""),
    ("nontarget_median_across_norm", "Median-normalized (_norm)",
     "Across-batch background after scaling.",
     "median over all other proteins of their target_median_norm",
     "All other proteins, all batches.",
     "Recomputed, not rescaled: this background mixes batches with different factors."),
    ("enrichment_across_norm", "Median-normalized (_norm)",
     "Across-batch enrichment after scaling.",
     "target_median_norm / nontarget_median_across_norm", "As above.", ""),
    ("pvalue_across_norm", "Median-normalized (_norm)",
     "Across-batch p-value after scaling.",
     "Mann-Whitney U, two-sided: normalized replicates vs normalized background medians",
     "As above.", "Forced to 1.0 when n_detected = 0."),
    ("label_across_norm", "Median-normalized (_norm)",
     "Hit call on the normalized across-batch statistics.",
     "1 if enrichment_across_norm >= 5 AND pvalue_across_norm <= 0.05, else 0",
     "As above.", ""),

    ("target_median_pctnorm", "Percentile-normalized (_pctnorm)",
     "This row's median intensity after quantile normalization.",
     "Per batch: value -> its rank percentile within the batch -> the value at that "
     "percentile of the POOLED (all-batch) distribution of replicates. "
     "target_median_pctnorm = median of the transformed replicates.",
     "This row; the map is shared by the whole batch.",
     "Forces every batch onto one common intensity distribution."),
    ("nontarget_median_within_pctnorm", "Percentile-normalized (_pctnorm)",
     "Within-batch background on the percentile scale.",
     "median over the other same-batch proteins of their target_median_pctnorm",
     "Same-batch proteins.", ""),
    ("enrichment_within_pctnorm", "Percentile-normalized (_pctnorm)",
     "Within-batch enrichment on the percentile scale.",
     "target_median_pctnorm / nontarget_median_within_pctnorm",
     "Same-batch proteins.", ""),
    ("pvalue_within_pctnorm", "Percentile-normalized (_pctnorm)",
     "Within-batch p-value on the percentile scale.",
     "Mann-Whitney U, two-sided: transformed replicates vs transformed background medians",
     "Same-batch proteins.",
     "Computed on the transformed values, NOT copied from pvalue_within: the transform "
     "is not strictly increasing (it flattens low values onto the floor), so it creates "
     "ties and changes the ranks."),
    ("label_within_pctnorm", "Percentile-normalized (_pctnorm)",
     "Hit call on the percentile within-batch statistics.",
     "1 if enrichment_within_pctnorm >= 5 AND pvalue_within_pctnorm <= 0.05, else 0",
     "Same-batch proteins.", ""),
    ("nontarget_median_across_pctnorm", "Percentile-normalized (_pctnorm)",
     "Across-batch background on the percentile scale.",
     "median over all other proteins of their target_median_pctnorm",
     "All other proteins, all batches.", ""),
    ("enrichment_across_pctnorm", "Percentile-normalized (_pctnorm)",
     "Across-batch enrichment on the percentile scale.",
     "target_median_pctnorm / nontarget_median_across_pctnorm", "As above.", ""),
    ("pvalue_across_pctnorm", "Percentile-normalized (_pctnorm)",
     "Across-batch p-value on the percentile scale.",
     "Mann-Whitney U, two-sided: transformed replicates vs transformed background medians",
     "As above.", "Forced to 1.0 when n_detected = 0."),
    ("label_across_pctnorm", "Percentile-normalized (_pctnorm)",
     "Hit call on the percentile across-batch statistics.",
     "1 if enrichment_across_pctnorm >= 5 AND pvalue_across_pctnorm <= 0.05, else 0",
     "As above.", ""),

    ("bead_signal", "Bead specificity",
     "How strongly this compound sticks to the beads alone (no protein).",
     "max( median(5-bead replicates), median(10-bead replicates) )",
     "Bead reference workbook, matched on COMPOUND_ID.",
     "NaN when the compound is absent from the bead run."),
    ("bead_detected", "Bead specificity",
     "Whether the compound was seen on beads at all.",
     "1 if bead_signal > 3000, else 0", "As above.", ""),
    ("bead_ratio", "Bead specificity",
     "Protein signal relative to the bead-only signal. Low = the protein adds little "
     "over the beads.",
     "bead_ratio = target_median / bead_signal", "As above.", ""),
    ("bead_binder_flag_r5", "Bead specificity",
     "Strict non-specific-binder flag.",
     "1 if bead_signal > 3000 AND bead_ratio < 5, else 0", "As above.",
     "Candidate false positive: mostly bead binding."),
    ("bead_binder_flag_r10", "Bead specificity",
     "Lenient non-specific-binder flag.",
     "1 if bead_signal > 3000 AND bead_ratio < 10, else 0", "As above.",
     "Wider net than the r5 flag."),
]

MATRIX = [
    ("COMPOUND_ID", "One row per molecule that is an ASMS hit (BINARY_LABEL = 1) on at "
                    "least one protein.", ""),
    ("SMILES", "Structure of the molecule.", ""),
    ("PROTEIN_NAME", "The protein(s) the molecule was called a hit on, comma-joined in "
                     "the same order as the analysis columns.", ""),
    ("n_floor_3000_of_median", "How many proteins show no signal for this molecule. "
                               "This is the row sort key (descending).",
     "count of proteins whose median intensity <= 3000 (not-measured counts as floor)"),
    ("n_floor_3000_in_replicas", "The same count at replicate level.",
     "count of individual replicate cells <= 3000, out of n_proteins x 3"),
    ("<all 35 added columns>", "Taken from the file(s) where this molecule is a hit; "
                               "comma-joined when it is a hit on several proteins, "
                               "aligned with PROTEIN_NAME.", ""),
    ("<protein> (median)", "Median ASMS intensity of the molecule on that protein; "
                           "not-measured is filled with the 3000 floor so the table "
                           "matches the heatmap.", "median(REP1, REP2, REP3)"),
    ("<protein>_REP1/2/3", "The raw replicate intensities behind each median cell.", ""),
]

PARAMS = [
    ("Floor", "3000", "Instrument reporting floor. A value at 3000 means 'not detected'. "
                      "Missing replicates are set to the floor for every column except "
                      "fold_articlebased_within."),
    ("Enrichment threshold", "5", "Used by every label_* column (paper's threshold)."),
    ("p-value threshold", "0.05", "Used by every label_* column."),
    ("Strong signal", ">= 1e6", "Replicate intensity cut-off used by signal_flag."),
    ("Moderate signal", "1e4 to 1e6", "Replicate intensity band used by signal_flag."),
    ("Within batch", "same ASMS_BATCH_NAME",
     "The other proteins run alongside this one in the same batch - 7 of them when the "
     "batch is complete. This is the control set used by the paper."),
    ("Across batches", "all other proteins",
     "Every other protein in the input folder, all batches pooled."),
    ("A (article formula)", "mean(REP1, REP2, REP3)",
     "Peak intensity used by fold_articlebased_within only."),
    ("M (everything else)", "median(REP1, REP2, REP3)",
     "Peak intensity used by all median-based columns; more robust to one bad replicate."),
    ("Statistical test", "Mann-Whitney U, two-sided",
     "Rank-based, so it does not assume normal intensities. Compares this row's 3 "
     "replicates against the background proteins' median values."),
    ("Missing replicates", "two conventions",
     "Median / normalized / bead columns treat a missing replicate as the 3000 floor "
     "('not detected'). fold_articlebased_within instead drops the protein from the "
     "background (so N shrinks) and is NaN for the unmeasured row, matching the "
     "pipeline's own fold_change column."),
    ("Source", "multibatchcodes/run_multibatch_pipeline.py",
     "Single entry point; regenerate this workbook with "
     "multibatchcodes/make_column_dictionary.py."),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
GROUP_FILLS = {
    "Article equation 1": "FCE4D6",
    "Per molecule": "FFF2CC",
    "Median enrichment - within batch": "DDEBF7",
    "Median enrichment - across batches": "E2EFDA",
    "Median-normalized (_norm)": "EDEDED",
    "Percentile-normalized (_pctnorm)": "F2E7F5",
    "Bead specificity": "D9D9D9",
}


def style(ws, widths, wrap_from=1, fill_col=None):
    """Bold header row, frozen panes, wrapped text, per-group row colours."""
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for j, cell in enumerate(row, start=1):
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(j >= wrap_from))
        if fill_col is not None:
            colour = GROUP_FILLS.get(row[fill_col].value)
            if colour:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=colour)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", default=DEFAULT_OUT)
    a = ap.parse_args()
    os.makedirs(os.path.dirname(a.out) or ".", exist_ok=True)

    added = pd.DataFrame(ADDED, columns=[
        "Column", "Group", "What it means", "Formula", "Background / scope", "Notes"])
    matrix = pd.DataFrame(MATRIX, columns=["Column", "What it means", "Formula"])
    params = pd.DataFrame(PARAMS, columns=["Setting", "Value", "Meaning"])

    with pd.ExcelWriter(a.out, engine="openpyxl") as xw:
        added.to_excel(xw, sheet_name="Added columns", index=False)
        matrix.to_excel(xw, sheet_name="Matrix export", index=False)
        params.to_excel(xw, sheet_name="Parameters", index=False)
        style(xw.sheets["Added columns"], [34, 34, 52, 60, 40, 56],
              wrap_from=3, fill_col=1)
        style(xw.sheets["Matrix export"], [30, 62, 52], wrap_from=2)
        style(xw.sheets["Parameters"], [24, 26, 72], wrap_from=3)

    print(f"saved {a.out}")
    print(f"  Added columns : {len(added)} rows")
    print(f"  Matrix export : {len(matrix)} rows")
    print(f"  Parameters    : {len(params)} rows")


if __name__ == "__main__":
    main()
