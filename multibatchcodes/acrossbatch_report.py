# -*- coding: utf-8 -*-
"""Across-batch selection report: one Excel sheet per protein.

Reads a *_modified folder produced by run_multibatch_pipeline.py and, FOR EACH
PROTEIN independently:

  1. sorts by each of the five across-batch metrics and takes the top N (200)
        enrich_across_medscale_medbg      enrich_across_quantscale_medbg
        enrich_across_medscale_meanbg     enrich_across_quantscale_meanbg
        Normalized score1
     -> up to 5 x 200 = 1000 rows, then MERGED to unique compounds. `selected_by`
        records which metrics picked each row and `n_metrics` how many, so a
        compound found by all five is easy to spot.
  2. splits the survivors into three blocks, written in this order:

        A. NEW CANDIDATES      no fill    tag = the replicate signal profile,
                                          e.g. "3 strong", "2 moderate + 1 weak"
                                          sorted by Normalized score1, best first
        B. FLOOR / BEAD        light blue tag = "has floor" | "bead binder" |
                                          "has floor and bead binder"
        C. ALREADY SELECTED    gray       tag = "Selected before based on within
                                          batch fold number"  (BINARY_LABEL = 1),
                                          kept at the BOTTOM

     A row qualifies for C first (it was already a hit), then B, then A -- so the
     blocks are disjoint and every selected row appears exactly once.

  3. Signal profile uses the pipeline's own thresholds:
        strong   >= 1e6
        moderate >= 1e4 and < 1e6
        weak     >  3000 and < 1e4
     Block A cannot contain a floored replicate by construction (those go to B).

Generic over datasets -- point --dir at any *_modified folder:

    python multibatchcodes/acrossbatch_report.py --dir MultiBatch_20to22_modified
    python multibatchcodes/acrossbatch_report.py --dir MultiBatch_20to36_modified --top 200

Output: <dataset>_acrossbatch_output.xlsx, one sheet per protein, plus a SUMMARY
sheet counting the blocks per protein.
"""

import os
import re
import glob
import argparse
import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ---- thresholds, kept identical to run_multibatch_pipeline.py ---------------
FLOOR = 3000.0
STRONG_MIN = 1e6
MODERATE_MIN = 1e4
REP = ["POS_INT_REP1", "POS_INT_REP2", "POS_INT_REP3"]

METRICS = [
    "enrich_across_medscale_medbg",
    "enrich_across_medscale_meanbg",
    "enrich_across_quantscale_medbg",
    "enrich_across_quantscale_meanbg",
    "Normalized score1",
]
SHORT = {                                   # compact names for `selected_by`
    "enrich_across_medscale_medbg": "med/med",
    "enrich_across_medscale_meanbg": "med/mean",
    "enrich_across_quantscale_medbg": "quant/med",
    "enrich_across_quantscale_meanbg": "quant/mean",
    "Normalized score1": "NS1",
}

OUT_ROOT = "MultiBatch_AcrossBatchDecisions"    # <root>/<dataset>/…

FILL_GRAY = PatternFill("solid", fgColor="D9D9D9")
FILL_BLUE = PatternFill("solid", fgColor="DDEBF7")
FILL_HEAD = PatternFill("solid", fgColor="404040")

TAG_SELECTED = "Selected before based on within batch fold number"

# columns to carry into the report, in order; silently skipped if absent
REPORT_COLS = [
    "tag", "block", "selected_by", "n_metrics",
    "COMPOUND_ID", "ASMS_BATCH_NAME", "SMILES",
    *REP, "n_detected", "target_median", "signal_flag",
    "fold_change", "enrichment_within", "ns1_enrichment_within",
    *METRICS,
    "bead_signal", "bead_detected", "bead_ratio",
    "BINARY_LABEL", "SPR_LABEL", "SPR_CATEGORY",
    "MZ", "RT (min)",
]


def signal_profile(row_reps):
    """Describe the three replicate intensities, e.g. '2 strong + 1 moderate'."""
    v = np.asarray(row_reps, dtype=float)
    v = v[np.isfinite(v)]
    n_strong = int((v >= STRONG_MIN).sum())
    n_mod = int(((v >= MODERATE_MIN) & (v < STRONG_MIN)).sum())
    n_weak = int(((v > FLOOR) & (v < MODERATE_MIN)).sum())
    n_floor = int((v <= FLOOR).sum())
    parts = []
    for k, name in ((n_strong, "strong"), (n_mod, "moderate"),
                    (n_weak, "weak"), (n_floor, "floor")):
        if k:
            parts.append(f"{k} {name}")
    return " + ".join(parts) if parts else "none"


def floor_bead_tag(row_reps, bead_detected):
    """'has floor' / 'bead binder' / 'has floor and bead binder' / ''."""
    v = np.asarray(row_reps, dtype=float)
    has_floor = bool(np.any(~np.isfinite(v)) or np.any(v <= FLOOR))
    is_bead = bool(pd.notna(bead_detected) and float(bead_detected) == 1)
    if has_floor and is_bead:
        return "has floor and bead binder"
    if has_floor:
        return "has floor"
    if is_bead:
        return "bead binder"
    return ""


def eligible(df, min_detected):
    """Rows allowed into the ranking at all.

    A row whose three replicates are ALL at the floor is not a measurement, so it
    must never compete for a top-N slot. This is not cosmetic: the per-batch
    scaling multiplies the floor itself, so an undetected compound scores

        sgcto_21  3000 x 1.0905 = 3271.64  ->  E ~ 1.09
        sgcto_20  3000 x 1.0000 = 3000.00  ->  E ~ 1.00
        sgcto_22  3000 x 0.4966 = 1489.85  ->  E ~ 0.50

    i.e. the COLD batch's non-detections outrank the hot batch's non-detections,
    purely from the scale factor. Before this filter, 1,275 of 11,657 selected
    rows on 20to22 were all-floored -- 298 of BRD1's 567 -- essentially all of
    them in sgcto_21, and none at all in sgcto_22. Pure scaling artifact.

    Partially floored rows (1-2 replicates at the floor) still compete; they are
    tagged "has floor" in block B, which is what the tagging is for.
    """
    R = df[REP].to_numpy(dtype=float)
    n_det = np.nansum(R > FLOOR, axis=1)
    return n_det >= min_detected


def build_protein_sheet(df, top_n, min_detected=1):
    """Select, merge and classify one protein's rows. Returns the report frame."""
    have = [m for m in METRICS if m in df.columns]
    if not have:
        return None

    df = df.loc[eligible(df, min_detected)]
    if df.empty:
        return None

    picked = {}                                    # positional index -> [metrics]
    for m in have:
        s = pd.to_numeric(df[m], errors="coerce")
        idx = s.nlargest(min(top_n, s.notna().sum())).index
        for i in idx:
            picked.setdefault(i, []).append(SHORT.get(m, m))
    if not picked:
        return None

    out = df.loc[list(picked.keys())].copy()
    out["selected_by"] = [", ".join(picked[i]) for i in out.index]
    out["n_metrics"] = [len(picked[i]) for i in out.index]

    reps = out[REP].to_numpy(dtype=float)
    bead = (out["bead_detected"] if "bead_detected" in out.columns
            else pd.Series(np.nan, index=out.index))
    binlab = pd.to_numeric(out.get("BINARY_LABEL", pd.Series(np.nan, index=out.index)),
                           errors="coerce")

    fb = [floor_bead_tag(reps[k], bead.iloc[k]) for k in range(len(out))]
    is_prev = (binlab == 1).to_numpy()
    is_fb = np.array([bool(t) for t in fb]) & ~is_prev
    is_new = ~is_prev & ~is_fb

    tag = np.empty(len(out), dtype=object)
    block = np.empty(len(out), dtype=object)
    tag[is_prev] = TAG_SELECTED
    block[is_prev] = "C_already_selected"
    for k in np.where(is_fb)[0]:
        tag[k] = fb[k]
    block[is_fb] = "B_floor_or_bead"
    for k in np.where(is_new)[0]:
        tag[k] = signal_profile(reps[k])
    block[is_new] = "A_new_candidate"
    out["tag"], out["block"] = tag, block

    ns1 = pd.to_numeric(out.get("Normalized score1",
                                pd.Series(np.nan, index=out.index)), errors="coerce")
    out["_ns1"] = ns1
    # A first (best Normalized score1 on top), then B, then C at the bottom
    out = out.sort_values(["block", "_ns1"], ascending=[True, False]).drop(columns="_ns1")

    cols = [c for c in REPORT_COLS if c in out.columns]
    return out[cols].reset_index(drop=True)


def style_sheet(ws, frame):
    """Header, per-block fills, freeze panes, column widths."""
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = FILL_HEAD
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 30

    blocks = frame["block"].to_numpy() if "block" in frame.columns else []
    for r, b in enumerate(blocks, start=2):
        fill = (FILL_GRAY if b.startswith("C") else
                FILL_BLUE if b.startswith("B") else None)
        if fill is None:
            continue
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill

    for c, name in enumerate(frame.columns, start=1):
        width = 46 if name in ("SMILES", "tag") else \
                22 if name in ("selected_by", "SPR_CATEGORY", "signal_flag") else \
                max(11, min(20, len(str(name)) + 3))
        ws.column_dimensions[get_column_letter(c)].width = width


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--dir", required=True,
                    help="a *_modified folder written by run_multibatch_pipeline.py")
    ap.add_argument("--top", type=int, default=200,
                    help="rows to take from each metric (default 200)")
    ap.add_argument("--root", default=OUT_ROOT,
                    help=f"top output folder (default {OUT_ROOT}/). A subfolder is "
                         "made per dataset, e.g. 20to22/ and 20to36/.")
    ap.add_argument("--no-split", action="store_true",
                    help="skip the one-file-per-protein copies, write only the "
                         "combined workbook")
    ap.add_argument("--min-detected", type=int, default=0,
                    help="a row needs at least this many replicates ABOVE the 3000 "
                         "floor to compete for a top-N slot. DEFAULT 0 = no filter, "
                         "all-floored rows compete and are tagged 'has floor' in "
                         "block B. Pass 1 to exclude them -- on 20to22 that removed "
                         "1,275 all-floored rows and surfaced only 68 extra "
                         "candidates, so it is off by default.")
    a = ap.parse_args()

    src = a.dir.rstrip("/\\")
    tag = os.path.basename(src).replace("_modified", "")      # MultiBatch_20to22
    short = re.sub(r"^MultiBatch[_-]?", "", tag) or tag       # 20to22
    out_dir = os.path.join(a.root, short)                     # .../20to22
    split_dir = os.path.join(out_dir, "per_protein")
    os.makedirs(out_dir, exist_ok=True)
    if not a.no_split:
        os.makedirs(split_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"{tag}_acrossbatch_output.xlsx")
    paths = sorted(glob.glob(os.path.join(src, "*.csv")))
    if not paths:
        raise SystemExit(f"no CSV files in {src}/")

    first = pd.read_csv(paths[0], nrows=0).columns.tolist()
    missing = [m for m in METRICS if m not in first]
    if missing:
        raise SystemExit(
            f"{src} is missing {len(missing)} across-batch column(s): {missing}\n"
            f"Re-run: python multibatchcodes/run_multibatch_pipeline.py "
            f"--input-dir {tag}")

    print(f"input : {src}/ ({len(paths)} proteins)\ntop-N : {a.top} per metric "
          f"x {len(METRICS)} metrics\ncombined: {out_path}"
          + ("" if a.no_split else f"\nper protein: {split_dir}/<protein>.xlsx") + "\n")

    summary = []
    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        for p in paths:
            prot = os.path.splitext(os.path.basename(p))[0]
            df = pd.read_csv(p, low_memory=False)
            rep = build_protein_sheet(df, a.top, a.min_detected)
            if rep is None or rep.empty:
                print(f"  {prot:<12} no rows selected -- skipped")
                continue
            # Excel sheet names: <=31 chars, no []:*?/\
            sheet = re.sub(r"[\[\]:*?/\\]", "_", prot)[:31]
            rep.to_excel(xl, sheet_name=sheet, index=False)
            style_sheet(xl.book[sheet], rep)

            if not a.no_split:      # the same sheet again, as its own workbook
                one = os.path.join(split_dir, f"{sheet}.xlsx")
                with pd.ExcelWriter(one, engine="openpyxl") as x1:
                    rep.to_excel(x1, sheet_name=sheet, index=False)
                    style_sheet(x1.book[sheet], rep)
            n = rep["block"].value_counts()
            summary.append({"protein": prot, "rows_in_report": len(rep),
                            "A_new_candidate": int(n.get("A_new_candidate", 0)),
                            "B_floor_or_bead": int(n.get("B_floor_or_bead", 0)),
                            "C_already_selected": int(n.get("C_already_selected", 0)),
                            "picked_by_all_5": int((rep["n_metrics"] == 5).sum())})
            print(f"  {prot:<12} {len(rep):>5} rows  "
                  f"(A {summary[-1]['A_new_candidate']:>4} | "
                  f"B {summary[-1]['B_floor_or_bead']:>4} | "
                  f"C {summary[-1]['C_already_selected']:>4})")

        s = pd.DataFrame(summary)
        s.to_excel(xl, sheet_name="SUMMARY", index=False)
        style_sheet(xl.book["SUMMARY"], s.assign(block=""))
        xl.book.move_sheet("SUMMARY", offset=-len(xl.book.sheetnames) + 1)

    print(f"\nTOTAL  A {s.A_new_candidate.sum():,} | B {s.B_floor_or_bead.sum():,} "
          f"| C {s.C_already_selected.sum():,}")
    print(f"Saved {out_path}")
    if not a.no_split:
        print(f"Saved {len(summary)} per-protein workbooks to {split_dir}/")


if __name__ == "__main__":
    main()
