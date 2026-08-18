"""
Full molecule x protein matrix (CSV) matching the floor-sorted heatmap
(hit_matrix_intensity_sorted_floor.py), with extra columns.

Source: MultiBatch_20to36_modified1/  (raw columns + computed enrichment columns)
Rows  : the 1647 molecules with BINARY_LABEL=1 for >=1 protein.
Cols  : all 136 proteins.

To match the heatmap, not-measured cells are filled with the 3000 floor.

Column layout (left to right):
  COMPOUND_ID, SMILES
  n_floor_3000_of_median      - # proteins where median(REP1-3) <= 3000  (SORT KEY, desc)
  n_floor_3000_in_replicas    - # individual replicate cells <= 3000 (out of 136*3 = 408)
  enrichment_within_batch     - from this molecule's hit (primary) protein
  pvalue_within_batch
  enrichment_across_batches   - across all 135 other proteins
  pvalue_across_batches
  enrichment_ref_protein      - which protein the 4 enrichment values came from
  <protein> (median)   x136   - median ASMS intensity per protein (missing -> 3000)
  <protein>_REP1/2/3   x408   - all raw replicate values per protein (missing -> 3000)

Rows sorted by n_floor_3000_of_median descending (most 3000s on top), same as heatmap.
"""
import csv, glob, os, statistics
from collections import defaultdict
import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill

SRC  = r"d:/0000-UHN/EASMS-data-processing/EASMS-data-processing/MultiBatch_20to36_modified1"
RAW  = r"d:/0000-UHN/EASMS-data-processing/EASMS-data-processing/MultiBatch_20to36"
OUT  = r"d:/0000-UHN/EASMS-data-processing/EASMS-data-processing/MultiBatch_20to36_plots"
os.makedirs(OUT, exist_ok=True)
CSV_OUT  = os.path.join(OUT, "MultiBatch_20to36_matrix_sorted_floor_full.csv")
XLSX_OUT = os.path.join(OUT, "MultiBatch_20to36_matrix_sorted_floor_full.xlsx")
REP_COLS = ['POS_INT_REP1', 'POS_INT_REP2', 'POS_INT_REP3']
FLOOR = 3000.0

def fnum(s):
    try: return float(s)
    except: return None

files = sorted(glob.glob(os.path.join(SRC, "*.csv")))
proteins = [os.path.basename(f)[:-4] for f in files]
print(f"proteins: {len(proteins)} | source: {SRC}")

# ---- pass 1: hit compounds, SMILES, per-protein hit count ----
hit_compounds = set()
smiles = {}
prot_hits = defaultdict(int)
for f in files:
    prot = os.path.basename(f)[:-4]
    r = csv.reader(open(f, encoding="utf-8", errors="replace")); hdr = next(r)
    I = {c: hdr.index(c) for c in ['COMPOUND_ID', 'SMILES', 'BINARY_LABEL'] if c in hdr}
    for row in r:
        comp = row[I['COMPOUND_ID']].strip()
        if 'SMILES' in I and comp not in smiles:
            smiles[comp] = row[I['SMILES']].strip()
        if row[I['BINARY_LABEL']].strip() == '1':
            hit_compounds.add(comp); prot_hits[prot] += 1
print(f"hit compounds: {len(hit_compounds)}")

# protein column order: by hit count desc (same as heatmap x-axis)
prot_order = sorted(proteins, key=lambda p: -prot_hits[p])
prank = {p: i for i, p in enumerate(prot_order)}

# PROTEIN_NAME per protein file, read from the RAW MultiBatch_20to36 folder
protein_name = {}
for p in proteins:
    rf = os.path.join(RAW, p + ".csv")
    try:
        r = csv.reader(open(rf, encoding="utf-8", errors="replace")); hdr = next(r)
        if 'PROTEIN_NAME' in hdr:
            j = hdr.index('PROTEIN_NAME')
            row = next(r)              # constant per file -> first row is enough
            protein_name[p] = row[j].strip()
    except StopIteration:
        pass
    protein_name.setdefault(p, p)

# ---- the appended analysis columns to grab: BY -> end of modified1 (target_median .. bead flags)
# Everything modified1 added on top of the raw MultiBatch_20to36 columns. 'within' is here.
raw_hdr = next(csv.reader(open(sorted(glob.glob(os.path.join(RAW, "*.csv")))[0],
                               encoding="utf-8", errors="replace")))
mod_hdr = next(csv.reader(open(files[0], encoding="utf-8", errors="replace")))
EXTRA_COLS = mod_hdr[len(raw_hdr):]      # BY .. DF
print(f"appended columns grabbed (BY..end): {len(EXTRA_COLS)} -> {EXTRA_COLS[0]} .. {EXTRA_COLS[-1]}")

# ---- pass 2: per (compound, protein) median, 3 replicas, all appended cols PER HIT FILE ----
# For each molecule we collect the appended analysis columns from EVERY protein file where
# it is a hit (BINARY_LABEL=1) -- the files the molecule was selected on. A molecule hit in
# several files yields several aligned values (comma-joined later, in protein-rank order).
median_v = defaultdict(dict)              # comp -> {prot: median}
reps_v   = defaultdict(dict)              # comp -> {prot: [r1, r2, r3]}
hits_of  = defaultdict(list)             # comp -> [(rank, prot, {extra col -> raw str}), ...]
for f in files:
    prot = os.path.basename(f)[:-4]
    r = csv.reader(open(f, encoding="utf-8", errors="replace")); hdr = next(r)
    I = {c: hdr.index(c) for c in
         ['COMPOUND_ID', 'BINARY_LABEL'] + REP_COLS + EXTRA_COLS if c in hdr}
    for row in r:
        comp = row[I['COMPOUND_ID']].strip()
        if comp not in hit_compounds:
            continue
        reps = [fnum(row[I[c]]) if c in I else None for c in REP_COLS]
        reps_v[comp][prot] = reps
        good = [x for x in reps if x is not None]
        if good:
            median_v[comp][prot] = statistics.median(good)
        # record all appended columns from EACH file where this molecule is a hit (label=1)
        if row[I['BINARY_LABEL']].strip() == '1':
            vals = {c: (row[I[c]].strip() if c in I and I[c] < len(row) else '')
                    for c in EXTRA_COLS}
            hits_of[comp].append((prank[prot], prot, vals))

# ---- build matrices; missing -> FLOOR (to match heatmap) ----
compounds = sorted(hit_compounds)

# median matrix
Mmed = np.array([[median_v.get(c, {}).get(p, np.nan) for p in prot_order]
                 for c in compounds], dtype=float)
n_missing_med = int((~np.isfinite(Mmed)).sum())
Mmed[~np.isfinite(Mmed)] = FLOOR

# replica matrix (136*3), missing/non-numeric -> FLOOR
rep_headers = [f"{p}_{rc.replace('POS_INT_','')}" for p in prot_order for rc in REP_COLS]
def rep_triplet(c, p):
    t = reps_v.get(c, {}).get(p)
    if t is None:
        return [FLOOR, FLOOR, FLOOR]
    return [x if (x is not None) else FLOOR for x in t]
Mrep = np.array([[v for p in prot_order for v in rep_triplet(c, p)]
                 for c in compounds], dtype=float)

# ---- floor counts ----
n_floor_median = (Mmed <= FLOOR).sum(axis=1)     # out of 136
n_floor_reps   = (Mrep <= FLOOR).sum(axis=1)     # out of 136*3 = 408

# ---- sort rows by n_floor_3000_of_median desc (most floors on top) ----
order = np.argsort(-n_floor_median, kind='stable')
compounds      = [compounds[i] for i in order]
Mmed           = Mmed[order]
Mrep           = Mrep[order]
n_floor_median = n_floor_median[order]
n_floor_reps   = n_floor_reps[order]
print(f"missing->floor(median)={n_missing_med}; "
      f"median floor-count top={n_floor_median[0]}..bottom={n_floor_median[-1]} (of {len(prot_order)}); "
      f"replica floor-count top={n_floor_reps[0]}..bottom={n_floor_reps[-1]} (of {len(prot_order)*3})")

# ---- build the aligned, comma-separated hit-file columns ----
# For each molecule, list its hit files ordered by protein hit-count rank. PROTEIN_NAME and
# every appended analysis column (BY..end) are joined in the SAME order, so position i of
# each list refers to the same hit protein.
pname_col = []
extra_cols = {c: [] for c in EXTRA_COLS}
for c in compounds:
    hits = sorted(hits_of.get(c, []), key=lambda t: t[0])   # by protein rank
    pname_col.append(", ".join(protein_name.get(p, p) for _, p, _ in hits))
    for ec in EXTRA_COLS:
        extra_cols[ec].append(", ".join(v.get(ec, '') for _, _, v in hits))

# ---- assemble the dataframe ----
meta = {
    "COMPOUND_ID": compounds,
    "SMILES": [smiles.get(c, "") for c in compounds],
    "PROTEIN_NAME": pname_col,
    "n_floor_3000_of_median": n_floor_median,
    "n_floor_3000_in_replicas": n_floor_reps,
}
meta.update({ec: extra_cols[ec] for ec in EXTRA_COLS})   # all BY..end appended columns
left = pd.DataFrame(meta)
median_headers = [f"{p} (median)" for p in prot_order]
med_cols = pd.DataFrame(Mmed, columns=median_headers)
rep_cols = pd.DataFrame(Mrep, columns=rep_headers)
df = pd.concat([left.reset_index(drop=True), med_cols, rep_cols], axis=1)

# ---- column order: base | enrichment+pvalue | signal_flag | bead | other | median | replicas
base_cols = ['COMPOUND_ID', 'SMILES', 'PROTEIN_NAME',
             'n_floor_3000_of_median', 'n_floor_3000_in_replicas']
enr_p_cols = [c for c in EXTRA_COLS if c.startswith('enrichment') or c.startswith('pvalue')]
bead_cols  = [c for c in EXTRA_COLS if c.startswith('bead')]
placed = set(enr_p_cols) | {'signal_flag'} | set(bead_cols)
other_cols = [c for c in EXTRA_COLS if c not in placed]     # target_median, nontarget_*, label_*, norm_*
col_order = (base_cols + enr_p_cols + ['signal_flag'] + bead_cols + other_cols
             + median_headers + rep_headers)
df = df[col_order]

df.to_csv(CSV_OUT, index=False)
print(f"saved {CSV_OUT}")

# ---- column fills: bead=gray, median=light blue, replicas=light green ----
GRAY  = PatternFill('solid', fgColor='D9D9D9')
BLUE  = PatternFill('solid', fgColor='DDEBF7')
GREEN = PatternFill('solid', fgColor='E2EFDA')
fill_of = {}
for c in bead_cols:       fill_of[c] = GRAY
for c in median_headers:  fill_of[c] = BLUE
for c in rep_headers:     fill_of[c] = GREEN

xlsx_path = XLSX_OUT
for attempt in range(5):
    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as xw:
            df.to_excel(xw, sheet_name="matrix_sorted_floor", index=False)
            ws = xw.sheets["matrix_sorted_floor"]
            nrow = df.shape[0]
            for j, cname in enumerate(df.columns, start=1):   # 1-based columns
                fill = fill_of.get(cname)
                if fill is None:
                    continue
                for i in range(1, nrow + 2):                  # header (1) + data rows
                    ws.cell(row=i, column=j).fill = fill
        break
    except PermissionError:
        xlsx_path = XLSX_OUT.replace(".xlsx", f"_{attempt + 2}.xlsx")
        print(f"  (locked - retrying as {os.path.basename(xlsx_path)})")
print(f"saved {xlsx_path}")
n_meta = df.shape[1] - len(prot_order) - len(rep_headers)
print(f"  shape: {df.shape[0]} molecules x {df.shape[1]} columns "
      f"({len(prot_order)} median + {len(rep_headers)} replica + {n_meta} meta)")
multi = sum(1 for c in compounds if len(hits_of.get(c, [])) > 1)
print(f"  molecules hit in >1 file (comma-separated enrichment): {multi}")
